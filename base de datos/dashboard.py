import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import os
import sqlite3
import mysql.connector
import urllib3
import requests
from bs4 import BeautifulSoup
import re
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import traceback
import threading
import certifi

# Importaciones desde conexion
from conexion import obtener_conexion, registrar_accion, hash_password, generar_id_cliente, obtener_ruta_recurso 
from chat_red import VentanaChat

TIPOS_EMPRESA = [
    "Industria manufacturera",
    "Comercio mayorista",
    "Comercio minorista / Retail",
    "Servicios profesionales",
    "Tecnología / Software",
    "Construcción",
    "Agricultura / Ganadería / Pesca",
    "Alimentos y bebidas",
    "Transporte / Logística",
    "Salud",
    "Educación",
    "Turismo / Hotelería",
    "Inmobiliario",
    "Finanzas / Seguros",
    "Telecomunicaciones",
    "Energía / Petróleo / Gas",
    "Minería",
    "Automotriz",
    "Textil / Moda",
    "Farmacéutica",
    "Medios / Publicidad",
    "Entretenimiento / Eventos",
    "Servicios personales",
    "ONG / Fundación",
    "Otro / No clasificado",
]

from chat_config import CHAT_SERVER_URL

# Nuevas importaciones para la Etapa 10 (Gráficos integrados en Tkinter)
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


# ================= AUTOMATIZACIÓN DE LA TASA BCV =================
def obtener_tasa_bcv_automatica():
    """Obtiene la tasa oficial del USD directamente desde el sitio del BCV."""
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    try:
        url = "https://www.bcv.org.ve/"
        response = requests.get(url, headers=headers, verify=certifi.where(), timeout=8)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            div_dolar = soup.find('div', id='dolar')
            if div_dolar:
                texto = div_dolar.get_text().strip()
                match = re.search(r'\d+,\d+', texto)
                if match:
                    return float(match.group().replace(',', '.'))
    except Exception:
        pass
    try:
        api_url = "https://open.er-api.com/v6/latest/USD"
        res = requests.get(api_url, timeout=5).json()
        if "rates" in res and "VES" in res["rates"]:
            return float(res["rates"]["VES"])
    except Exception:
        pass
    return None


class DashboardSGC:
    def __init__(self, root):
        self.root = root
        self.root.title("Macilitano Consulting Group C.A. · Sistema de Gestión")
        self.root.geometry("1180x740")
        self.root.minsize(1020, 650)

        # Identidad visual de la aplicación.
        try:
            ruta_icono_png = obtener_ruta_recurso("logo.png")
            if os.path.exists(ruta_icono_png):
                self._app_icon_photo = tk.PhotoImage(file=ruta_icono_png)
                self.root.iconphoto(True, self._app_icon_photo)
        except Exception:
            pass

        try:
            ruta_icono_ico = obtener_ruta_recurso("logo.ico")
            if os.path.exists(ruta_icono_ico):
                self.root.iconbitmap(ruta_icono_ico)
        except Exception:
            pass

        self.colores = {
            "sidebar": "#6b1426",       
            "botones_menu": "#1f1f1f",  
            "fondo_main": "#fdfaf2",    
            "oro": "#d4af37",           
            "rojo_eliminar": "#9e2a2b", 
            "verde_aprobar": "#2a9d8f",
            "blanco_tarjeta": "#ffffff",
            "texto_oscuro": "#333333"
        }
        
        self.estilo = ttk.Style()
        self.estilo.theme_use("default")
        
        self.estilo.configure("Treeview.Heading", 
                    background=self.colores["botones_menu"], 
                    foreground=self.colores["oro"], 
                    font=("Arial", 10, "bold"), 
                    relief="flat")
        self.estilo.map("Treeview.Heading", background=[('active', '#2a2a2a')])
        
        self.estilo.configure("Treeview", 
                background="white", 
                fieldbackground="white", 
                foreground="black", 
                font=("Arial", 10), 
                rowheight=24)
        self.estilo.map("Treeview", 
                background=[('selected', self.colores["sidebar"])], 
                foreground=[('selected', 'white')])

        # Acabado global de controles ttk.
        self.estilo.configure(
            "Treeview",
            background="white",
            fieldbackground="white",
            foreground=self.colores["texto_oscuro"],
            font=("Segoe UI", 9),
            rowheight=30,
            borderwidth=0,
        )
        self.estilo.configure(
            "Treeview.Heading",
            background=self.colores["botones_menu"],
            foreground=self.colores["oro"],
            font=("Segoe UI", 9, "bold"),
            padding=(8, 8),
            relief="flat",
        )
        self.estilo.configure("TCombobox", padding=6, font=("Segoe UI", 9))
        self.estilo.configure("Vertical.TScrollbar", arrowsize=13)
        self.estilo.configure("Horizontal.TScrollbar", arrowsize=13)
        
        self.ultimo_mensaje_notificado = 0
        self.notificaciones_chat = 0
        self.solicitudes_token = None
        self.ultima_solicitud_web_id = 0
        self.monitor_solicitudes_activo = False
        self.monitor_solicitudes_en_consulta = False
        self.ids_solicitudes_vistas = set()
        self.monitor_solicitudes_inicializado = False
        self.total_solicitudes_pendientes = 0

        # Estado de sesión y referencias opcionales declaradas explícitamente.
        # Esto conserva el comportamiento actual y permite al analizador
        # estático conocer estos atributos desde la construcción del dashboard.
        self.usuario_autenticado: str = ""
        self.rol_autenticado: str = ""
        self.password_autenticado: str | None = ""
        self.ventana_chat: VentanaChat | None = None

        self.iniciar_monitor_mensajes()
        self.configurar_interfaz()

    def configurar_interfaz(self):
        sidebar = tk.Frame(
            self.root,
            bg=self.colores["sidebar"],
            width=240
        )
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        marca = tk.Frame(sidebar, bg=self.colores["sidebar"])
        marca.pack(fill="x", pady=(18, 12))

        ruta_logo = obtener_ruta_recurso("logo.png")
        if os.path.exists(ruta_logo):
            try:
                self._tk_logo = tk.PhotoImage(file=ruta_logo)
                ancho_logo = max(1, self._tk_logo.width())
                factor = max(1, round(ancho_logo / 150))
                if factor > 1:
                    self._tk_logo = self._tk_logo.subsample(factor, factor)

                lbl_logo = tk.Label(
                    marca,
                    image=self._tk_logo,
                    bg=self.colores["sidebar"],
                    bd=0
                )
                lbl_logo.pack()
            except Exception:
                tk.Label(
                    marca,
                    text="MACILITANO",
                    fg="white",
                    bg=self.colores["sidebar"],
                    font=("Georgia", 15, "bold")
                ).pack()
        else:
            tk.Label(
                marca,
                text="MACILITANO",
                fg="white",
                bg=self.colores["sidebar"],
                font=("Georgia", 15, "bold")
            ).pack()

        tk.Label(
            marca,
            text="CONSULTING GROUP C.A.",
            fg=self.colores["oro"],
            bg=self.colores["sidebar"],
            font=("Segoe UI", 8, "bold")
        ).pack(pady=(7, 0))

        tk.Frame(
            sidebar,
            bg=self.colores["oro"],
            height=1
        ).pack(fill="x", padx=18, pady=(0, 10))

        # Construir botones del menú y guardar referencias para control de permisos
        self.menu_buttons = {}
        self.notas_pendientes = 0

        def add_menu_button(texto, comando):
            btn = tk.Button(
                sidebar,
                text=texto,
                bg=self.colores["botones_menu"],
                fg=self.colores["oro"],
                font=("Segoe UI", 10, "bold"),
                relief="flat",
                bd=0,
                activebackground=self.colores["oro"],
                activeforeground=self.colores["botones_menu"],
                anchor="w",
                padx=18,
                pady=8,
                cursor="hand2",
                command=comando
            )
            btn.pack(fill="x", pady=3, padx=12)
            self.menu_buttons[texto] = btn
            return btn

        add_menu_button("Inicio", self.mostrar_inicio)
        add_menu_button("Clientes", self.mostrar_clientes)
        add_menu_button("Asesores", self.mostrar_asesores)
        add_menu_button("Servicios", self.mostrar_servicios)
        add_menu_button("Gestión de Costos", self.mostrar_gestion_costos)
        self.btn_solicitudes = add_menu_button("Solicitudes", self.mostrar_notificaciones)
        
        self.btn_chat = add_menu_button(
    "💬 Chats",
    self.abrir_chat
    )
    
        # Crear el botón de Historial pero su visibilidad/estado dependerá del rol
        self.btn_historial = add_menu_button("Historial Logs", self.mostrar_historial)

        self.area_principal = tk.Frame(self.root, bg=self.colores["fondo_main"])
        self.area_principal.pack(side="right", fill="both", expand=True)
        
        self.mostrar_inicio()
        self.actualizar_alertas_globales()

        # main.py asigna usuario/rol justo después de crear el Dashboard.
        # Aplicamos permisos un instante después para leer el rol real.
        try:
            self.root.after(250, self.aplicar_permisos)
        except Exception:
            pass
    
        # ============================================================
    # CHAT INTERNO
    # ============================================================

        # ============================================================
    # CHAT INTERNO
    # ============================================================

    def abrir_chat(self):

        usuario_actual = getattr(
            self,
            "usuario_autenticado",
            None
        )
        password_actual = getattr(
            self,
            "password_autenticado",
            None
        )

        if not usuario_actual:

            self._mensaje_corporativo(
                "Chat no disponible",
                "No se pudo identificar al usuario actual.",
                tipo="advertencia"
            )

            return

        try:

            ventana = VentanaChat(
                self.root,
                usuario_actual,
                password_actual
            )

            ventana.transient(
                self.root
            )

            # Guardamos referencia para evitar
            # que Python la destruya
            self.ventana_chat = ventana

            ventana.protocol(
                "WM_DELETE_WINDOW",
                lambda: self.cerrar_ventana_chat()
            )

        except Exception as e:

            self._mensaje_corporativo(
                "No se pudo abrir el chat",
                "Ocurrió un problema al iniciar el módulo de conversaciones.",
                tipo="error",
                detalle=str(e)
            )

    def cerrar_ventana_chat(self):

        ventana_chat = self.ventana_chat
        if ventana_chat is not None:
            try:
                ventana_chat.cerrar()
            except Exception:
                pass

            self.ventana_chat = None
    def _rol_actual(self):
        return str(getattr(self, "rol_autenticado", "") or "").strip().lower()

    def _es_administrador(self):
        return self._rol_actual() == "administrador"

    def _requiere_administrador(self, accion="realizar esta acción"):
        if self._es_administrador():
            return True

        self._mensaje_corporativo(
            "Acceso restringido",
            f"Solo un Administrador puede {accion}.",
            tipo="advertencia"
        )
        return False

    def aplicar_permisos(self):
        """Aplica la matriz de permisos de la aplicación según el rol autenticado."""
        es_admin = self._es_administrador()

        # Menús exclusivamente administrativos.
        for nombre in ("Asesores", "Historial Logs"):
            btn = getattr(self, "menu_buttons", {}).get(nombre)
            if not btn:
                continue

            if es_admin:
                if not btn.winfo_manager():
                    btn.pack(fill="x", pady=4, padx=10)
                btn.config(state="normal")
            else:
                try:
                    btn.pack_forget()
                except Exception:
                    btn.config(state="disabled")

        # Compatibilidad con la referencia histórica del botón de logs.
        if hasattr(self, "btn_historial"):
            if es_admin:
                if not self.btn_historial.winfo_manager():
                    self.btn_historial.pack(fill="x", pady=4, padx=10)
                self.btn_historial.config(state="normal")
            else:
                try:
                    self.btn_historial.pack_forget()
                except Exception:
                    self.btn_historial.config(state="disabled")

    # ============================================================
    # VENTANAS / DIÁLOGOS CORPORATIVOS
    # ============================================================
    def _centrar_ventana(self, ventana, ancho, alto):
        try:
            self.root.update_idletasks()
            x = self.root.winfo_x() + max(0, (self.root.winfo_width() - ancho) // 2)
            y = self.root.winfo_y() + max(0, (self.root.winfo_height() - alto) // 2)
            ventana.geometry(f"{ancho}x{alto}+{x}+{y}")
        except Exception:
            ventana.geometry(f"{ancho}x{alto}")

    def _crear_modal_corporativo(self, titulo, subtitulo="", ancho=520, alto=320, icono="◆"):
        ventana = tk.Toplevel(self.root)
        ventana.title(titulo)
        ventana.configure(bg=self.colores["fondo_main"])

        try:
            if hasattr(self, "_app_icon_photo"):
                ventana.iconphoto(False, self._app_icon_photo)
        except Exception:
            pass

        try:
            ruta_icono_ico = obtener_ruta_recurso("logo.ico")
            if os.path.exists(ruta_icono_ico):
                ventana.iconbitmap(ruta_icono_ico)
        except Exception:
            pass
        ventana.transient(self.root)
        ventana.grab_set()
        ventana.resizable(False, False)
        self._centrar_ventana(ventana, ancho, alto)

        cabecera = tk.Frame(ventana, bg=self.colores["sidebar"], height=82)
        cabecera.pack(fill="x")
        cabecera.pack_propagate(False)

        tk.Label(
            cabecera, text=icono, bg=self.colores["sidebar"],
            fg=self.colores["oro"], font=("Arial", 22, "bold")
        ).pack(side="left", padx=(22, 12), pady=20)

        textos = tk.Frame(cabecera, bg=self.colores["sidebar"])
        textos.pack(side="left", fill="y", pady=16)

        tk.Label(
            textos, text=titulo, bg=self.colores["sidebar"], fg="white",
            font=("Arial", 14, "bold")
        ).pack(anchor="w")

        if subtitulo:
            tk.Label(
                textos, text=subtitulo, bg=self.colores["sidebar"],
                fg="#eadba7", font=("Arial", 9)
            ).pack(anchor="w", pady=(3, 0))

        cuerpo = tk.Frame(ventana, bg=self.colores["fondo_main"])
        cuerpo.pack(fill="both", expand=True, padx=24, pady=20)
        return ventana, cuerpo

    def _dialogo_texto_corporativo(
        self, titulo, mensaje, valor_inicial="", etiqueta="", ancho=540, ocultar=False
    ):
        ventana, cuerpo = self._crear_modal_corporativo(
            titulo, "Complete la información solicitada", ancho=ancho, alto=290, icono="✦"
        )

        tk.Label(
            cuerpo, text=mensaje, bg=self.colores["fondo_main"],
            fg=self.colores["texto_oscuro"], font=("Arial", 10),
            justify="left", wraplength=ancho - 70
        ).pack(anchor="w", pady=(2, 10))

        if etiqueta:
            tk.Label(
                cuerpo, text=etiqueta, bg=self.colores["fondo_main"],
                fg=self.colores["sidebar"], font=("Arial", 9, "bold")
            ).pack(anchor="w", pady=(2, 5))

        marco = tk.Frame(cuerpo, bg="white", highlightbackground="#d9d1c7", highlightthickness=1)
        marco.pack(fill="x", pady=(0, 18))

        entrada = tk.Entry(
            marco, font=("Arial", 11), relief="flat", bd=0,
            show="*" if ocultar else ""
        )
        entrada.pack(fill="x", padx=12, pady=10)
        if valor_inicial:
            entrada.insert(0, valor_inicial)
            entrada.select_range(0, tk.END)

        resultado: dict[str, str | None] = {"valor": None}

        def confirmar(event=None):
            valor = entrada.get().strip()
            if not valor:
                return
            resultado["valor"] = valor
            ventana.destroy()

        def cancelar(event=None):
            ventana.destroy()

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x")

        tk.Button(
            botones, text="Cancelar", bg=self.colores["botones_menu"],
            fg=self.colores["oro"], font=("Arial", 10, "bold"),
            relief="flat", padx=18, pady=7, command=cancelar
        ).pack(side="right")

        tk.Button(
            botones, text="Confirmar", bg=self.colores["sidebar"], fg="white",
            font=("Arial", 10, "bold"), relief="flat", padx=18, pady=7,
            command=confirmar
        ).pack(side="right", padx=(0, 8))

        entrada.bind("<Return>", confirmar)
        entrada.bind("<Escape>", cancelar)
        entrada.focus_set()
        ventana.wait_window()
        return resultado["valor"]

    def _mensaje_corporativo(self, titulo, mensaje, tipo="info", detalle=""):
        iconos = {
            "info": ("i", self.colores["sidebar"]),
            "exito": ("✓", self.colores["verde_aprobar"]),
            "advertencia": ("!", "#b7791f"),
            "error": ("×", self.colores["rojo_eliminar"]),
        }
        icono, color = iconos.get(tipo, iconos["info"])

        ventana, cuerpo = self._crear_modal_corporativo(
            titulo, "", ancho=520, alto=310 if detalle else 270, icono=icono
        )

        tk.Label(
            cuerpo, text=mensaje, bg=self.colores["fondo_main"],
            fg=self.colores["texto_oscuro"], font=("Arial", 11, "bold"),
            justify="left", wraplength=450
        ).pack(anchor="w", pady=(5, 8))

        if detalle:
            caja = tk.Frame(cuerpo, bg="white", highlightbackground="#ddd5c8", highlightthickness=1)
            caja.pack(fill="x", pady=(2, 14))
            tk.Label(
                caja, text=detalle, bg="white", fg=self.colores["texto_oscuro"],
                font=("Arial", 9), justify="left", wraplength=430
            ).pack(anchor="w", padx=12, pady=10)

        tk.Button(
            cuerpo, text="Aceptar", bg=color, fg="white",
            font=("Arial", 10, "bold"), relief="flat",
            padx=22, pady=7, command=ventana.destroy
        ).pack(anchor="e", pady=(10, 0))

        ventana.bind("<Return>", lambda e: ventana.destroy())
        ventana.bind("<Escape>", lambda e: ventana.destroy())
        ventana.wait_window()

    def _confirmar_corporativo(
        self,
        titulo,
        mensaje,
        detalle="",
        texto_confirmar="Confirmar",
        peligro=False,
    ):
        """Confirmación consistente para operaciones importantes."""
        ventana, cuerpo = self._crear_modal_corporativo(
            titulo,
            "Revise la información antes de continuar",
            ancho=560,
            alto=340 if detalle else 290,
            icono="!" if peligro else "?"
        )

        tk.Label(
            cuerpo,
            text=mensaje,
            bg=self.colores["fondo_main"],
            fg=self.colores["texto_oscuro"],
            font=("Arial", 11, "bold"),
            justify="left",
            wraplength=485,
        ).pack(anchor="w", pady=(4, 10))

        if detalle:
            caja = tk.Frame(
                cuerpo,
                bg="white",
                highlightbackground="#ddd5c8",
                highlightthickness=1,
            )
            caja.pack(fill="x", pady=(0, 16))

            tk.Label(
                caja,
                text=detalle,
                bg="white",
                fg=self.colores["texto_oscuro"],
                font=("Arial", 9),
                justify="left",
                wraplength=455,
            ).pack(anchor="w", padx=12, pady=10)

        resultado = {"valor": False}

        def aceptar():
            resultado["valor"] = True
            ventana.destroy()

        def cancelar():
            ventana.destroy()

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x", pady=(8, 0))

        tk.Button(
            botones,
            text="Cancelar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=18,
            pady=7,
            command=cancelar,
        ).pack(side="right")

        tk.Button(
            botones,
            text=texto_confirmar,
            bg=self.colores["rojo_eliminar"] if peligro else self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=18,
            pady=7,
            command=aceptar,
        ).pack(side="right", padx=(0, 8))

        ventana.bind("<Escape>", lambda e: cancelar())
        ventana.wait_window()
        return resultado["valor"]

    def crear_panel_principal(self, titulo):
        for widget in self.area_principal.winfo_children():
            widget.destroy()

        cabecera = tk.Frame(
            self.area_principal,
            bg=self.colores["fondo_main"]
        )
        cabecera.pack(fill="x", padx=28, pady=(20, 12))

        tk.Label(
            cabecera,
            text=titulo,
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Segoe UI", 20, "bold")
        ).pack(anchor="w")

        tk.Frame(
            cabecera,
            bg=self.colores["oro"],
            height=2
        ).pack(fill="x", pady=(9, 0))

        return self.area_principal

    def actualizar_alertas_globales(self):
        """Inicia el monitor de solicitudes sin bloquear la interfaz."""
        self.notas_pendientes = 0

        if not getattr(self, "monitor_solicitudes_activo", False):
            self.monitor_solicitudes_activo = True
            # Arranca rápido para reducir la ventana en la que una solicitud
            # recién llegada podría pasar desapercibida.
            self.root.after(350, self.monitor_solicitudes_web)

    def _api_solicitudes(self, metodo, ruta, **kwargs):
        """Cliente centralizado para solicitudes web alojadas en PythonAnywhere."""
        if not CHAT_SERVER_URL.strip():
            raise RuntimeError("CHAT_SERVER_URL no está configurado")

        if not self.solicitudes_token:
            usuario = getattr(self, "usuario_autenticado", None)
            password = getattr(self, "password_autenticado", None)
            if not usuario or not password:
                raise RuntimeError("No hay credenciales de sesión disponibles")

            login = requests.post(
                f"{CHAT_SERVER_URL.rstrip('/')}/api/session",
                json={"username": usuario, "password": password},
                timeout=10,
                verify=certifi.where(),
            )
            login.raise_for_status()
            data_login = login.json()

            if not data_login.get("ok"):
                raise RuntimeError(
                    data_login.get("error", "No se pudo iniciar sesión remota")
                )

            self.solicitudes_token = data_login["token"]

        headers = {"Authorization": f"Bearer {self.solicitudes_token}"}

        respuesta = requests.request(
            metodo,
            f"{CHAT_SERVER_URL.rstrip('/')}{ruta}",
            headers=headers,
            timeout=10,
            verify=certifi.where(),
            **kwargs,
        )

        if respuesta.status_code == 401:
            self.solicitudes_token = None
            return self._api_solicitudes(metodo, ruta, **kwargs)

        respuesta.raise_for_status()
        data = respuesta.json()

        if not data.get("ok"):
            raise RuntimeError(data.get("error", "Error en API de solicitudes"))

        return data

    def _actualizar_badge_solicitudes(self, cantidad):
        """Actualiza el texto del menú sin recrear botones."""
        self.total_solicitudes_pendientes = int(cantidad or 0)

        if not hasattr(self, "btn_solicitudes"):
            return

        if self.total_solicitudes_pendientes > 0:
            texto = f"Solicitudes   ● {self.total_solicitudes_pendientes}"
        else:
            texto = "Solicitudes"

        try:
            self.btn_solicitudes.config(text=texto)
        except Exception:
            pass

    def monitor_solicitudes_web(self):
        """
        Dispara una consulta en segundo plano.

        Antes la petición HTTPS se hacía directamente dentro del hilo de Tkinter.
        Si la red tardaba, también se atrasaban los siguientes callbacks y la alerta
        podía aparecer muchos segundos después. Ahora la interfaz nunca espera a la red.
        """
        if not getattr(self, "monitor_solicitudes_activo", False):
            return

        if getattr(self, "monitor_solicitudes_en_consulta", False):
            # Ya hay una consulta en curso: no lanzamos otra encima.
            try:
                self.root.after(1000, self.monitor_solicitudes_web)
            except Exception:
                pass
            return

        self.monitor_solicitudes_en_consulta = True

        hilo = threading.Thread(
            target=self._consultar_solicitudes_monitor,
            daemon=True,
        )
        hilo.start()

        # El siguiente ciclo se programa desde aquí y la red corre aparte.
        try:
            self.root.after(2000, self.monitor_solicitudes_web)
        except Exception:
            pass

    def _consultar_solicitudes_monitor(self):
        """Trabajo de red del monitor; nunca crea widgets desde este hilo."""
        try:
            data = self._api_solicitudes(
                "GET",
                "/api/solicitudes?estado=Pendiente",
            )
            items = data.get("items", [])

            # Entregamos el resultado al hilo principal de Tkinter.
            try:
                self.root.after(
                    0,
                    lambda datos=list(items): self._procesar_resultado_monitor(datos),
                )
            except Exception:
                self.monitor_solicitudes_en_consulta = False

        except Exception as e:
            print(f"Monitor solicitudes web: {e}")
            self.monitor_solicitudes_en_consulta = False

    def _procesar_resultado_monitor(self, items):
        """Procesa cambios y muestra alertas solamente desde Tkinter."""
        try:
            ids_actuales = {
                int(x.get("id", 0))
                for x in items
                if int(x.get("id", 0)) > 0
            }

            self._actualizar_badge_solicitudes(len(items))

            if not self.monitor_solicitudes_inicializado:
                # Primera fotografía. Registramos lo que ya existía al abrir
                # el sistema para no bombardear con solicitudes antiguas.
                self.ids_solicitudes_vistas = set(ids_actuales)
                self.ultima_solicitud_web_id = max(ids_actuales, default=0)
                self.monitor_solicitudes_inicializado = True
                return

            nuevas = [
                solicitud
                for solicitud in items
                if int(solicitud.get("id", 0)) not in self.ids_solicitudes_vistas
            ]

            for solicitud in sorted(
                nuevas,
                key=lambda x: int(x.get("id", 0)),
            ):
                solicitud_id = int(solicitud.get("id", 0))

                if solicitud_id:
                    self.ids_solicitudes_vistas.add(solicitud_id)
                    self.ultima_solicitud_web_id = max(
                        self.ultima_solicitud_web_id,
                        solicitud_id,
                    )

                self.mostrar_notificacion_solicitud(solicitud)

            # Conservamos IDs vistos aunque la solicitud después sea aprobada
            # o rechazada; así jamás se notifica dos veces la misma.
            self.ids_solicitudes_vistas.update(ids_actuales)

        finally:
            self.monitor_solicitudes_en_consulta = False

    def mostrar_notificacion_solicitud(self, solicitud):
        n = tk.Toplevel(self.root)
        n.title("Nueva solicitud")

        try:
            if hasattr(self, "_app_icon_photo"):
                n.iconphoto(False, self._app_icon_photo)
        except Exception:
            pass

        try:
            ruta_icono_ico = obtener_ruta_recurso("logo.ico")
            if os.path.exists(ruta_icono_ico):
                n.iconbitmap(ruta_icono_ico)
        except Exception:
            pass

        try:
            self.root.bell()
        except Exception:
            pass
        n.resizable(False, False)
        n.configure(bg="white")
        try:
            n.attributes("-topmost", True)
            n.after(2500, lambda: n.winfo_exists() and n.attributes("-topmost", False))
        except Exception:
            pass

        ancho, alto = 430, 205
        n.update_idletasks()
        sw, sh = n.winfo_screenwidth(), n.winfo_screenheight()
        n.geometry(f"{ancho}x{alto}+{sw-ancho-28}+{max(25, sh-alto-75)}")

        cabecera = tk.Frame(n, bg=self.colores["sidebar"], height=55)
        cabecera.pack(fill="x")
        cabecera.pack_propagate(False)

        tk.Label(
            cabecera, text="NUEVA SOLICITUD WEB", bg=self.colores["sidebar"],
            fg="white", font=("Arial", 10, "bold")
        ).pack(side="left", padx=16, pady=17)

        tk.Label(
            cabecera, text="●", bg=self.colores["sidebar"],
            fg=self.colores["oro"], font=("Arial", 12, "bold")
        ).pack(side="right", padx=16)

        cuerpo = tk.Frame(n, bg="white")
        cuerpo.pack(fill="both", expand=True, padx=16, pady=12)

        cliente = solicitud.get("cliente_potencial") or "Cliente"
        servicio = solicitud.get("servicio_interes") or "Servicio no especificado"

        tk.Label(
            cuerpo, text=cliente, bg="white", fg=self.colores["sidebar"],
            font=("Arial", 12, "bold")
        ).pack(anchor="w")

        tk.Label(
            cuerpo, text=servicio, bg="white", fg=self.colores["texto_oscuro"],
            font=("Arial", 9), wraplength=390, justify="left"
        ).pack(anchor="w", pady=(3, 10))

        botones = tk.Frame(cuerpo, bg="white")
        botones.pack(fill="x")

        tk.Button(
            botones, text="Ver solicitudes", bg=self.colores["sidebar"], fg="white",
            font=("Arial", 9, "bold"), relief="flat", padx=14, pady=6,
            command=lambda: (n.destroy(), self.mostrar_notificaciones())
        ).pack(side="right")

        tk.Button(
            botones, text="Cerrar", bg="#f2eee8", fg=self.colores["texto_oscuro"],
            font=("Arial", 9), relief="flat", padx=12, pady=6,
            command=n.destroy
        ).pack(side="right", padx=(0, 7))

        try:
            n.after(10000, lambda: n.winfo_exists() and n.destroy())
        except Exception:
            pass

    def mostrar_inicio(self):
        panel = self.crear_panel_principal("Panel de Control Ejecutivo")

        # ============================================================
        # MÉTRICAS REALES DEL SISTEMA
        # ============================================================
        total_clientes_activos = 0
        total_servicios = 0
        facturado_mes = 0.0
        cobrado_mes = 0.0
        por_cobrar = 0.0
        solicitudes_pendientes = 0
        clientes_recientes = []
        operaciones_recientes = []
        meses_grafico = []
        ingresos_grafico = []

        try:
            conn = obtener_conexion()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT COUNT(*)
                FROM clientes
                WHERE LOWER(COALESCE(estado, 'Activo')) = 'activo'
            """)
            total_clientes_activos = int(cursor.fetchone()[0] or 0)

            cursor.execute("SELECT COUNT(*) FROM servicios")
            total_servicios = int(cursor.fetchone()[0] or 0)

            cursor.execute("""
                SELECT COALESCE(SUM(total_usd), 0)
                FROM solicitudes_servicio
                WHERE strftime('%Y-%m', fecha_solicitud) = strftime('%Y-%m', 'now', 'localtime')
            """)
            facturado_mes = float(cursor.fetchone()[0] or 0)

            cursor.execute("""
                SELECT COALESCE(SUM(total_usd), 0)
                FROM solicitudes_servicio
                WHERE COALESCE(estado_pago, 'Pendiente') = 'Pagado'
                  AND strftime('%Y-%m', COALESCE(fecha_pago, fecha_solicitud))
                      = strftime('%Y-%m', 'now', 'localtime')
            """)
            cobrado_mes = float(cursor.fetchone()[0] or 0)

            cursor.execute("""
                SELECT COALESCE(SUM(total_usd), 0)
                FROM solicitudes_servicio
                WHERE COALESCE(estado_pago, 'Pendiente') = 'Pendiente'
            """)
            por_cobrar = float(cursor.fetchone()[0] or 0)

            cursor.execute("""
                SELECT codigo, nombre, servicio, estado
                FROM clientes
                ORDER BY id DESC
                LIMIT 5
            """)
            clientes_recientes = cursor.fetchall()

            cursor.execute("""
                SELECT
                    c.nombre,
                    s.nombre_servicio,
                    ss.total_usd,
                    COALESCE(ss.estado_pago, 'Pendiente'),
                    COALESCE(ss.fecha_solicitud, '')
                FROM solicitudes_servicio ss
                INNER JOIN clientes c ON c.id = ss.cliente_id
                INNER JOIN servicios s ON s.id = ss.servicio_id
                ORDER BY ss.id DESC
                LIMIT 5
            """)
            operaciones_recientes = cursor.fetchall()

            # Últimos 6 meses reales de facturación.
            cursor.execute("""
                WITH RECURSIVE meses(n) AS (
                    SELECT 5
                    UNION ALL
                    SELECT n - 1 FROM meses WHERE n > 0
                )
                SELECT
                    strftime('%Y-%m', date('now', 'start of month', printf('-%d months', n))) AS ym,
                    COALESCE((
                        SELECT SUM(ss.total_usd)
                        FROM solicitudes_servicio ss
                        WHERE strftime('%Y-%m', ss.fecha_solicitud) =
                              strftime('%Y-%m', date('now', 'start of month', printf('-%d months', n)))
                    ), 0) AS total
                FROM meses
                ORDER BY ym
            """)

            filas_meses = cursor.fetchall()

            nombres_meses = {
                "01": "Ene", "02": "Feb", "03": "Mar", "04": "Abr",
                "05": "May", "06": "Jun", "07": "Jul", "08": "Ago",
                "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dic"
            }

            for ym, total in filas_meses:
                partes = str(ym).split("-")
                etiqueta = nombres_meses.get(partes[1], partes[1]) if len(partes) == 2 else str(ym)
                meses_grafico.append(etiqueta)
                ingresos_grafico.append(float(total or 0))

            conn.close()

        except Exception as e:
            print(f"Error consultando métricas reales del panel: {e}")

        # Solicitudes web pendientes. Inicio no falla si el servidor no responde.
        try:
            data_web = self._api_solicitudes("GET", "/api/solicitudes")
            for solicitud in data_web.get("items", []):
                estado_web = str(
                    solicitud.get("estado") or "Pendiente"
                ).strip().lower()

                if estado_web in {"pendiente", "contactado"}:
                    solicitudes_pendientes += 1
        except Exception:
            pass

        tasa_bcv = obtener_tasa_bcv_automatica() or 0.0

        # ============================================================
        # TARJETAS EJECUTIVAS
        # ============================================================
        frame_tarjetas = tk.Frame(panel, bg=self.colores["fondo_main"])
        frame_tarjetas.pack(fill="x", padx=25, pady=(8, 10))

        metricas = [
            ("Clientes Activos", str(total_clientes_activos)),
            ("Servicios", str(total_servicios)),
            ("Solicitudes Pend.", str(solicitudes_pendientes)),
            ("Facturado Mes", f"$ {facturado_mes:,.2f}"),
            ("Por Cobrar", f"$ {por_cobrar:,.2f}"),
            (
                "Tasa Oficial BCV",
                f"{tasa_bcv:,.4f} Bs/USD" if tasa_bcv else "No disponible"
            ),
        ]

        for idx, (titulo, valor) in enumerate(metricas):
            tarjeta = tk.Frame(
                frame_tarjetas,
                bg=self.colores["blanco_tarjeta"],
                highlightbackground=self.colores["oro"],
                highlightthickness=1,
                bd=0
            )
            tarjeta.grid(row=0, column=idx, padx=5, pady=3, sticky="nsew")
            frame_tarjetas.grid_columnconfigure(idx, weight=1)

            tk.Label(
                tarjeta,
                text=titulo,
                fg="#777777",
                bg=self.colores["blanco_tarjeta"],
                font=("Arial", 9, "bold")
            ).pack(pady=(9, 2), padx=7)

            tk.Label(
                tarjeta,
                text=valor,
                fg=self.colores["sidebar"],
                bg=self.colores["blanco_tarjeta"],
                font=("Arial", 13, "bold")
            ).pack(pady=(0, 9), padx=7)

        # Línea informativa secundaria.
        info = tk.Frame(panel, bg=self.colores["fondo_main"])
        info.pack(fill="x", padx=30, pady=(0, 8))

        tk.Label(
            info,
            text=f"Cobrado este mes: $ {cobrado_mes:,.2f}",
            bg=self.colores["fondo_main"],
            fg=self.colores["texto_oscuro"],
            font=("Arial", 9, "bold")
        ).pack(side="left")


        # ============================================================
        # CUERPO: GRÁFICO + ACTIVIDAD RECIENTE
        # ============================================================
        cuerpo = tk.Frame(panel, bg=self.colores["fondo_main"])
        cuerpo.pack(fill="both", expand=True, padx=25, pady=(0, 15))
        cuerpo.grid_columnconfigure(0, weight=3)
        cuerpo.grid_columnconfigure(1, weight=2)
        cuerpo.grid_rowconfigure(0, weight=1)

        # -------- Gráfico real --------
        frame_grafico = tk.Frame(
            cuerpo,
            bg=self.colores["blanco_tarjeta"],
            highlightbackground="#d9d4c8",
            highlightthickness=1
        )
        frame_grafico.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        tk.Label(
            frame_grafico,
            text="📊 Facturación real de los últimos 6 meses (USD)",
            bg=self.colores["blanco_tarjeta"],
            fg=self.colores["sidebar"],
            font=("Arial", 11, "bold")
        ).pack(anchor="w", padx=15, pady=(10, 4))

        if not meses_grafico:
            meses_grafico = ["—"]
            ingresos_grafico = [0]

        figura = Figure(figsize=(6, 3), dpi=100)
        eje = figura.add_subplot(111)
        eje.plot(
            meses_grafico,
            ingresos_grafico,
            marker="o",
            color=self.colores["sidebar"],
            linewidth=2
        )
        eje.fill_between(
            meses_grafico,
            ingresos_grafico,
            color=self.colores["sidebar"],
            alpha=0.08
        )
        eje.set_ylabel("USD")
        eje.grid(True, linestyle="--", alpha=0.35)
        figura.tight_layout()

        canvas = FigureCanvasTkAgg(figura, master=frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # -------- Actividad reciente --------
        frame_actividad = tk.Frame(
            cuerpo,
            bg=self.colores["blanco_tarjeta"],
            highlightbackground="#d9d4c8",
            highlightthickness=1
        )
        frame_actividad.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        tk.Label(
            frame_actividad,
            text="Actividad reciente",
            bg=self.colores["blanco_tarjeta"],
            fg=self.colores["sidebar"],
            font=("Arial", 11, "bold")
        ).pack(anchor="w", padx=15, pady=(10, 6))

        tk.Label(
            frame_actividad,
            text="Últimos clientes",
            bg=self.colores["blanco_tarjeta"],
            fg=self.colores["texto_oscuro"],
            font=("Arial", 9, "bold")
        ).pack(anchor="w", padx=15, pady=(2, 3))

        if clientes_recientes:
            for codigo, nombre, servicio, estado_cliente in clientes_recientes:
                fila = tk.Frame(frame_actividad, bg=self.colores["blanco_tarjeta"])
                fila.pack(fill="x", padx=15, pady=2)

                tk.Label(
                    fila,
                    text=f"{codigo or '—'}  {nombre}",
                    bg=self.colores["blanco_tarjeta"],
                    fg=self.colores["texto_oscuro"],
                    font=("Arial", 8, "bold"),
                    anchor="w"
                ).pack(fill="x")

                tk.Label(
                    fila,
                    text=f"{servicio or 'Sin servicio'} · {estado_cliente or 'Activo'}",
                    bg=self.colores["blanco_tarjeta"],
                    fg="#777777",
                    font=("Arial", 8),
                    anchor="w"
                ).pack(fill="x")
        else:
            tk.Label(
                frame_actividad,
                text="No hay clientes registrados.",
                bg=self.colores["blanco_tarjeta"],
                fg="#777777",
                font=("Arial", 8)
            ).pack(anchor="w", padx=15)

        ttk.Separator(frame_actividad, orient="horizontal").pack(
            fill="x", padx=15, pady=8
        )

        tk.Label(
            frame_actividad,
            text="Últimas operaciones",
            bg=self.colores["blanco_tarjeta"],
            fg=self.colores["texto_oscuro"],
            font=("Arial", 9, "bold")
        ).pack(anchor="w", padx=15, pady=(0, 3))

        if operaciones_recientes:
            for cliente, servicio, total, estado, fecha in operaciones_recientes:
                fila = tk.Frame(frame_actividad, bg=self.colores["blanco_tarjeta"])
                fila.pack(fill="x", padx=15, pady=2)

                tk.Label(
                    fila,
                    text=f"{cliente} · $ {float(total or 0):,.2f}",
                    bg=self.colores["blanco_tarjeta"],
                    fg=self.colores["texto_oscuro"],
                    font=("Arial", 8, "bold"),
                    anchor="w"
                ).pack(fill="x")

                tk.Label(
                    fila,
                    text=f"{servicio} · {estado} · {fecha}",
                    bg=self.colores["blanco_tarjeta"],
                    fg="#777777",
                    font=("Arial", 8),
                    anchor="w",
                    wraplength=350,
                    justify="left"
                ).pack(fill="x")
        else:
            tk.Label(
                frame_actividad,
                text="No hay operaciones registradas.",
                bg=self.colores["blanco_tarjeta"],
                fg="#777777",
                font=("Arial", 8)
            ).pack(anchor="w", padx=15)

    def mostrar_clientes(self):
        panel = self.crear_panel_principal("Gestión de Clientes")
        
        # Barra superior dividida en dos filas para evitar que los botones se solapen.
        # Fila 1: búsqueda y filtros.
        filtros_bar = tk.Frame(panel, bg=self.colores["fondo_main"])
        filtros_bar.pack(fill="x", padx=25, pady=(5, 5))

        lbl_buscar = tk.Label(
            filtros_bar,
            text="Nombre / empresa:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 10, "bold")
        )
        lbl_buscar.pack(side="left", padx=(0, 5))

        self.entry_buscar = tk.Entry(
            filtros_bar,
            font=("Arial", 10),
            width=22,
            relief="solid",
            bd=1
        )
        self.entry_buscar.pack(side="left", padx=5, ipady=3)
        self.entry_buscar.bind("<KeyRelease>", lambda event: self.actualizar_tabla_clientes())

        lbl_servicio = tk.Label(
            filtros_bar,
            text="Servicio:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 10, "bold")
        )
        lbl_servicio.pack(side="left", padx=(12, 5))

        self.combo_servicio = ttk.Combobox(
            filtros_bar,
            state="readonly",
            width=20
        )
        self.combo_servicio.pack(side="left", padx=5)
        self.combo_servicio.bind(
            "<<ComboboxSelected>>",
            lambda event: self.actualizar_tabla_clientes()
        )

        lbl_tipo = tk.Label(
            filtros_bar,
            text="Tipo cliente:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 10, "bold")
        )
        lbl_tipo.pack(side="left", padx=(12, 5))

        self.combo_tipo_cliente = ttk.Combobox(
            filtros_bar,
            state="readonly",
            width=22
        )
        self.combo_tipo_cliente.pack(side="left", padx=5)
        self.combo_tipo_cliente.bind(
            "<<ComboboxSelected>>",
            lambda event: self.actualizar_tabla_clientes()
        )

        try:
            conn = obtener_conexion()
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT nombre_servicio
                FROM servicios
                WHERE nombre_servicio IS NOT NULL
                  AND TRIM(nombre_servicio) != ''
                ORDER BY nombre_servicio
                """
            )
            servicios = [row[0] for row in cursor.fetchall()]

            cursor.execute(
                """
                SELECT DISTINCT servicio
                FROM clientes
                WHERE servicio IS NOT NULL
                  AND TRIM(servicio) != ''
                ORDER BY servicio
                """
            )
            for (servicio,) in cursor.fetchall():
                if servicio and servicio not in servicios:
                    servicios.append(servicio)

            tipos = ["Todos"] + TIPOS_EMPRESA

            cursor.execute(
                """
                SELECT DISTINCT industria
                FROM clientes
                WHERE industria IS NOT NULL
                  AND TRIM(industria) != ''
                ORDER BY industria
                """
            )
            for (tipo,) in cursor.fetchall():
                if tipo and tipo not in tipos:
                    tipos.append(tipo)

            conn.close()

        except Exception:
            servicios = []
            tipos = ["Todos"] + TIPOS_EMPRESA

        self.combo_servicio["values"] = ["Todos"] + servicios
        self.combo_servicio.set("Todos")

        self.combo_tipo_cliente["values"] = tipos
        self.combo_tipo_cliente.set("Todos")

        # Fila 2: acciones. "Agregar Cliente" queda separado y siempre visible.
        acciones_bar = tk.Frame(panel, bg=self.colores["fondo_main"])
        acciones_bar.pack(fill="x", padx=25, pady=(0, 10))

        btn_agregar = tk.Button(
            acciones_bar,
            text="➕ Agregar Cliente",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=14,
            pady=5,
            command=self.modal_agregar_cliente
        )
        btn_agregar.pack(side="left", padx=(0, 12))

        btn_ordenar = tk.Button(
            acciones_bar,
            text="↕ Ordenar columnas",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=10,
            command=self.ordenar_columnas_clientes
        )
        btn_ordenar.pack(side="right", padx=5)

        btn_pdf = tk.Button(
            acciones_bar,
            text="✅ Exportar PDF",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=10,
            command=self.exportar_clientes_pdf
        )
        btn_pdf.pack(side="right", padx=5)

        btn_excel = tk.Button(
            acciones_bar,
            text="✅ Exportar Excel",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=10,
            command=self.exportar_clientes_excel
        )
        btn_excel.pack(side="right", padx=5)

        btn_eliminar = tk.Button(
            acciones_bar,
            text="🗑 Eliminar",
            bg=self.colores["rojo_eliminar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=10,
            command=self.eliminar_cliente_bd
        )
        btn_eliminar.pack(side="right", padx=5)
        if not self._es_administrador():
            btn_eliminar.pack_forget()

        btn_editar = tk.Button(
            acciones_bar,
            text="✏ Editar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=10,
            command=self.modal_editar_cliente
        )
        btn_editar.pack(side="right", padx=5)

        tk.Label(panel, text="📋 Directorio de Contactos y Servicios", bg=self.colores["fondo_main"], 
                fg=self.colores["sidebar"], font=("Arial", 12, "bold")).pack(anchor="w", padx=25, pady=(5, 2))
        
        # NUEVO ORDEN DE COLUMNAS REQUERIDO
        columnas_c = (
            "ID", 
            "Código Cliente", 
            "Nombre Cliente", 
            "Servicio Solicitado", 
            "Teléfono de Contacto", 
            "Nombre Persona Contacto", 
            "Correo", 
            "Status"
        )
        self.tabla_clientes = ttk.Treeview(panel, columns=columnas_c, show="headings", height=15)
        
        # Configuración del ancho de las columnas ajustado al nuevo orden
        for col in columnas_c:
            self.tabla_clientes.heading(col, text=col)
            if col in ["Nombre Cliente", "Nombre Persona Contacto", "Correo"]:
                self.tabla_clientes.column(col, anchor="w", width=150)
            elif col in ["Servicio Solicitado"]:
                self.tabla_clientes.column(col, anchor="w", width=210)
            elif col in ["ID", "Código Cliente"]:
                self.tabla_clientes.column(col, anchor="center", width=80)
            else:
                self.tabla_clientes.column(col, anchor="center", width=110)
                
        self.tabla_clientes.pack(fill="both", expand=True, padx=25, pady=(0, 15))
        self.tabla_clientes.bind("<Double-1>", self.alternar_estado_servicio)
        self.actualizar_tabla_clientes()

    def exportar_clientes_excel(self):
        try:
            if not hasattr(self, 'tabla_clientes'):
                return
            path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx'), ('All Files', '*.*')])
            if not path:
                return
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.title = 'Clientes'

            columnas_raw = self.tabla_clientes['columns']
            columnas = (
                list(columnas_raw)
                if isinstance(columnas_raw, (tuple, list))
                else [columnas_raw]
            )
            ws.append(columnas)

            for item in self.tabla_clientes.get_children():
                valores_raw = self.tabla_clientes.item(item, 'values')
                valores = (
                    list(valores_raw)
                    if isinstance(valores_raw, (tuple, list))
                    else [valores_raw]
                )
                ws.append(valores)

            wb.save(path)
            self._mensaje_corporativo(
                "Exportación completada",
                "El directorio de clientes fue exportado correctamente a Excel.",
                tipo="exito",
                detalle=path
            )
        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo exportar",
                "Ocurrió un problema al generar el archivo Excel.",
                tipo="error",
                detalle=str(e)
            )

    def exportar_clientes_pdf(self):
        try:
            if not hasattr(self, 'tabla_clientes'):
                return
            path = filedialog.asksaveasfilename(defaultextension='.pdf', filetypes=[('PDF Files', '*.pdf'), ('All Files', '*.*')])
            if not path:
                return
            c = canvas.Canvas(path, pagesize=letter)
            width, height = letter
            columnas = list(self.tabla_clientes['columns'])
            x_positions = [50, 110, 200, 360, 520, 620, 720, 820]
            y = height - 50
            c.setFont('Helvetica-Bold', 10)
            for idx, col in enumerate(columnas):
                x = x_positions[idx] if idx < len(x_positions) else 50 + idx * 100
                c.drawString(x, y, str(col))
            c.setFont('Helvetica', 9)
            y -= 20
            for item in self.tabla_clientes.get_children():
                valores = self.tabla_clientes.item(item, 'values')
                if y < 50:
                    c.showPage()
                    y = height - 50
                    c.setFont('Helvetica-Bold', 10)
                    for idx, col in enumerate(columnas):
                        x = x_positions[idx] if idx < len(x_positions) else 50 + idx * 100
                        c.drawString(x, y, str(col))
                    c.setFont('Helvetica', 9)
                    y -= 20
                for idx, valor in enumerate(valores):
                    x = x_positions[idx] if idx < len(x_positions) else 50 + idx * 100
                    c.drawString(x, y, str(valor))
                y -= 18
            c.save()
            self._mensaje_corporativo(
                "Exportación completada",
                "El directorio de clientes fue exportado correctamente a PDF.",
                tipo="exito",
                detalle=path
            )
        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo exportar",
                "Ocurrió un problema al generar el archivo PDF.",
                tipo="error",
                detalle=str(e)
            )

    def ordenar_columnas_clientes(self):
        try:
            if not hasattr(self, 'tabla_clientes'):
                return
            columna = 'Nombre Cliente'
            orden = getattr(self, 'orden_cliente_asc', True)
            items = [(self.tabla_clientes.set(k, columna), k) for k in self.tabla_clientes.get_children('')]
            items.sort(key=lambda t: str(t[0]).lower(), reverse=not orden)
            for index, (_, k) in enumerate(items):
                self.tabla_clientes.move(k, '', index)
            self.orden_cliente_asc = not orden
            direccion = 'ascendente' if orden else 'descendente'
            self._mensaje_corporativo(
                "Tabla ordenada",
                f'La tabla fue ordenada por "{columna}" en orden {direccion}.',
                tipo="info"
            )
        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo ordenar",
                "Ocurrió un problema al reorganizar la tabla.",
                tipo="error",
                detalle=str(e)
            )

    def actualizar_tabla_clientes(self):
        """Actualiza el directorio de clientes aplicando filtros en tiempo real."""
        for item in self.tabla_clientes.get_children(): 
            self.tabla_clientes.delete(item)
            
        texto_busqueda = self.entry_buscar.get().strip() if hasattr(self, 'entry_buscar') else ""
        servicio_filtro = self.combo_servicio.get() if hasattr(self, 'combo_servicio') else "Todos"
        tipo_filtro = self.combo_tipo_cliente.get() if hasattr(self, 'combo_tipo_cliente') else "Todos"

        condiciones = []
        params = []

        if texto_busqueda:
            condiciones.append("(c.nombre LIKE ? OR c.correo LIKE ? OR c.codigo LIKE ? OR c.telefono LIKE ? OR COALESCE(c.nombre_contacto, c.nombre) LIKE ?)")
            texto_like = f"%{texto_busqueda}%"
            params.extend([texto_like, texto_like, texto_like, texto_like, texto_like])

        if servicio_filtro and servicio_filtro != "Todos":
            condiciones.append("COALESCE(s.nombre_servicio, c.servicio, '') LIKE ?")
            params.append(f"%{servicio_filtro}%")

        if tipo_filtro and tipo_filtro != "Todos":
            condiciones.append("LOWER(COALESCE(c.industria, '')) LIKE ?")
            params.append(f"%{tipo_filtro.lower()}%")

        where_sql = f" WHERE {' AND '.join(condiciones)}" if condiciones else ""

        try:
            conn = obtener_conexion()
            cursor = conn.cursor()
            
            cursor.execute("PRAGMA table_info(clientes)")
            columnas = [col[1] for col in cursor.fetchall()]
            if "estado" not in columnas:
                cursor.execute("ALTER TABLE clientes ADD COLUMN estado TEXT DEFAULT 'Activo'")
                conn.commit()

            # Estructuración de datos 
            # ID | Código Cliente | Nombre Cliente | Servicio Solicitado | Teléfono | Persona Contacto | Correo | Status
            cursor.execute(f"""
                SELECT 
                    c.id, 
                    c.codigo, 
                    c.nombre, 
                    IFNULL(COALESCE(s.nombre_servicio, c.servicio), 'Ninguno'), 
                    IFNULL(c.telefono, 'No registrado'),
                    IFNULL(c.nombre_contacto, c.nombre), 
                    IFNULL(c.correo, 'No registrado'), 
                    IFNULL(c.estado, 'Activo')
                FROM clientes c
                LEFT JOIN solicitudes_servicio ss ON c.id = ss.cliente_id
                LEFT JOIN servicios s ON ss.servicio_id = s.id
                {where_sql}
                GROUP BY c.id
            """, params)
            
            for fila in cursor.fetchall():
                self.tabla_clientes.insert("", "end", values=fila)
            conn.close()
        except Exception as e:
            print(f"Error al actualizar listas: {e}")


    # ----------------- GESTIÓN DE COSTOS / FINANZAS -----------------
    def _obtener_configuracion_financiera(self):
        conn = obtener_conexion()
        cur = conn.cursor()
        cur.execute("""
            SELECT iva_pct, igtf_pct,
                   igtf_efectivo, igtf_transferencia, igtf_pago_movil
            FROM configuracion_financiera
            WHERE id = 1
        """)
        fila = cur.fetchone()
        conn.close()

        if not fila:
            return {
                "iva_pct": 16.0,
                "igtf_pct": 3.0,
                "igtf_efectivo": 1,
                "igtf_transferencia": 0,
                "igtf_pago_movil": 0,
            }

        return {
            "iva_pct": float(fila[0]),
            "igtf_pct": float(fila[1]),
            "igtf_efectivo": int(fila[2]),
            "igtf_transferencia": int(fila[3]),
            "igtf_pago_movil": int(fila[4]),
        }

    def _aplica_igtf(self, metodo_pago, config=None):
        config = config or self._obtener_configuracion_financiera()
        metodo = (metodo_pago or "").strip().lower()

        if "efectivo" in metodo:
            return bool(config["igtf_efectivo"])
        if "transferencia" in metodo:
            return bool(config["igtf_transferencia"])
        if "pago movil" in metodo or "pago móvil" in metodo:
            return bool(config["igtf_pago_movil"])

        return False

    def configurar_parametros_financieros(self):

        if not self._requiere_administrador(
            "cambiar IVA, IGTF o sus reglas de aplicación"
        ):
            return

        try:
            cfg = self._obtener_configuracion_financiera()
        except Exception as e:
            self._mensaje_corporativo(
                "Configuración no disponible",
                "No fue posible cargar los parámetros financieros.",
                tipo="error",
                detalle=str(e),
            )
            return

        modal, cuerpo = self._crear_modal_corporativo(
            "Configuración financiera",
            "Parámetros aplicados a nuevas operaciones",
            ancho=620,
            alto=530,
            icono="⚙",
        )

        form = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        form.pack(fill="both", expand=True)
        form.grid_columnconfigure(1, weight=1)

        tk.Label(
            form,
            text="IVA (%)",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 10, "bold"),
        ).grid(row=0, column=0, sticky="w", padx=(0, 18), pady=10)

        marco_iva = tk.Frame(
            form,
            bg="white",
            highlightbackground="#d9d1c7",
            highlightthickness=1,
        )
        marco_iva.grid(row=0, column=1, sticky="ew", pady=8)

        ent_iva = tk.Entry(
            marco_iva,
            font=("Arial", 11),
            relief="flat",
            bd=0,
        )
        ent_iva.pack(fill="x", padx=10, pady=8)
        ent_iva.insert(0, str(cfg["iva_pct"]))

        tk.Label(
            form,
            text="IGTF (%)",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 10, "bold"),
        ).grid(row=1, column=0, sticky="w", padx=(0, 18), pady=10)

        marco_igtf = tk.Frame(
            form,
            bg="white",
            highlightbackground="#d9d1c7",
            highlightthickness=1,
        )
        marco_igtf.grid(row=1, column=1, sticky="ew", pady=8)

        ent_igtf = tk.Entry(
            marco_igtf,
            font=("Arial", 11),
            relief="flat",
            bd=0,
        )
        ent_igtf.pack(fill="x", padx=10, pady=8)
        ent_igtf.insert(0, str(cfg["igtf_pct"]))

        tk.Label(
            form,
            text="Aplicar IGTF en:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 10, "bold"),
        ).grid(
            row=2,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(20, 8),
        )

        var_efectivo = tk.BooleanVar(value=bool(cfg["igtf_efectivo"]))
        var_transferencia = tk.BooleanVar(
            value=bool(cfg["igtf_transferencia"])
        )
        var_pago_movil = tk.BooleanVar(
            value=bool(cfg["igtf_pago_movil"])
        )

        opciones = tk.Frame(
            form,
            bg="white",
            highlightbackground="#ddd5c8",
            highlightthickness=1,
        )
        opciones.grid(
            row=3,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(0, 8),
        )

        for texto_op, variable in [
            ("Efectivo", var_efectivo),
            ("Transferencia", var_transferencia),
            ("Pago Móvil", var_pago_movil),
        ]:
            tk.Checkbutton(
                opciones,
                text=texto_op,
                variable=variable,
                bg="white",
                activebackground="white",
                fg=self.colores["texto_oscuro"],
                selectcolor="white",
                font=("Arial", 10),
                padx=10,
                pady=7,
            ).pack(fill="x", anchor="w")

        tk.Label(
            form,
            text=(
                "Los cambios afectan nuevas operaciones. "
                "Los montos históricos registrados no se recalculan."
            ),
            bg=self.colores["fondo_main"],
            fg="#766b6d",
            font=("Arial", 9),
            wraplength=530,
            justify="left",
        ).grid(
            row=4,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(12, 0),
        )

        def guardar():
            try:
                iva = float(ent_iva.get().strip().replace(",", "."))
                igtf = float(ent_igtf.get().strip().replace(",", "."))

                if not (0 <= iva <= 100 and 0 <= igtf <= 100):
                    raise ValueError

            except ValueError:
                self._mensaje_corporativo(
                    "Valores inválidos",
                    "Las tasas deben ser números entre 0 y 100.",
                    tipo="advertencia",
                )
                return

            try:
                conn = obtener_conexion()
                cur = conn.cursor()
                cur.execute(
                    """
                    UPDATE configuracion_financiera
                    SET iva_pct = ?, igtf_pct = ?,
                        igtf_efectivo = ?, igtf_transferencia = ?,
                        igtf_pago_movil = ?
                    WHERE id = 1
                    """,
                    (
                        iva,
                        igtf,
                        1 if var_efectivo.get() else 0,
                        1 if var_transferencia.get() else 0,
                        1 if var_pago_movil.get() else 0,
                    ),
                )
                conn.commit()
                conn.close()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "ACTUALIZÓ CONFIGURACIÓN FINANCIERA",
                    f"IVA {iva:.2f}% | IGTF {igtf:.2f}%.",
                )

                modal.destroy()
                self.mostrar_gestion_costos()

                self._mensaje_corporativo(
                    "Configuración actualizada",
                    "Los nuevos parámetros financieros fueron guardados.",
                    tipo="exito",
                    detalle=f"IVA: {iva:.2f}%\nIGTF: {igtf:.2f}%",
                )

            except Exception as e:
                self._mensaje_corporativo(
                    "No se pudo guardar",
                    "Ocurrió un problema al actualizar la configuración.",
                    tipo="error",
                    detalle=str(e),
                )

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x", pady=(18, 0))

        tk.Button(
            botones,
            text="Cancelar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=modal.destroy,
        ).pack(side="right")

        tk.Button(
            botones,
            text="Guardar configuración",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=guardar,
        ).pack(side="right", padx=(0, 8))

        ent_iva.focus_set()

    def mostrar_gestion_costos(self):
        panel = self.crear_panel_principal("Gestión de Costos y Cuentas por Cobrar")

        try:
            cfg = self._obtener_configuracion_financiera()
        except Exception:
            cfg = {"iva_pct": 16.0, "igtf_pct": 3.0}

        # =========================
        # CABECERA
        # =========================
        cabecera = tk.Frame(panel, bg=self.colores["fondo_main"])
        cabecera.pack(fill="x", padx=25, pady=(0, 8))

        tk.Label(
            cabecera,
            text=f"IVA configurado: {cfg['iva_pct']:.2f}%   |   IGTF configurado: {cfg['igtf_pct']:.2f}%",
            bg=self.colores["fondo_main"],
            fg=self.colores["texto_oscuro"],
            font=("Arial", 10, "italic")
        ).pack(side="left")

        btn_config_tasas = tk.Button(
            cabecera,
            text="⚙ Configurar tasas",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=12,
            command=self.configurar_parametros_financieros
        )
        btn_config_tasas.pack(side="right")
        if not self._es_administrador():
            btn_config_tasas.pack_forget()

        # =========================
        # FILTROS
        # =========================
        filtros = tk.Frame(panel, bg=self.colores["fondo_main"])
        filtros.pack(fill="x", padx=25, pady=(0, 8))

        tk.Label(
            filtros, text="Cliente:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 9, "bold")
        ).pack(side="left")

        combo_cliente = ttk.Combobox(filtros, state="readonly", width=18)
        combo_cliente.pack(side="left", padx=(5, 10))

        tk.Label(
            filtros, text="Servicio:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 9, "bold")
        ).pack(side="left")

        combo_servicio = ttk.Combobox(filtros, state="readonly", width=20)
        combo_servicio.pack(side="left", padx=(5, 10))

        tk.Label(
            filtros, text="Estado:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 9, "bold")
        ).pack(side="left")

        combo_estado = ttk.Combobox(
            filtros,
            state="readonly",
            width=12,
            values=["Todos", "Pendiente", "Pagado"]
        )
        combo_estado.set("Todos")
        combo_estado.pack(side="left", padx=(5, 10))

        tk.Label(
            filtros, text="Desde:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 9, "bold")
        ).pack(side="left")

        entry_desde = tk.Entry(filtros, width=11)
        entry_desde.pack(side="left", padx=(5, 6))
        entry_desde.insert(0, "")

        tk.Label(
            filtros, text="Hasta:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 9, "bold")
        ).pack(side="left")

        entry_hasta = tk.Entry(filtros, width=11)
        entry_hasta.pack(side="left", padx=(5, 6))
        entry_hasta.insert(0, "")

        tk.Label(
            panel,
            text="Fechas opcionales en formato AAAA-MM-DD.",
            bg=self.colores["fondo_main"],
            fg="#777",
            font=("Arial", 8, "italic")
        ).pack(anchor="e", padx=30, pady=(0, 4))

        # =========================
        # TARJETAS / MÉTRICAS
        # =========================
        tarjetas = tk.Frame(panel, bg=self.colores["fondo_main"])
        tarjetas.pack(fill="x", padx=25, pady=(0, 10))

        valores_metricas = {}

        titulos = [
            ("Operaciones", "operaciones"),
            ("Facturado USD", "facturado"),
            ("Cobrado USD", "cobrado"),
            ("Por cobrar USD", "por_cobrar"),
        ]

        for idx, (titulo, clave) in enumerate(titulos):
            tarjeta = tk.Frame(
                tarjetas,
                bg=self.colores["blanco_tarjeta"],
                highlightbackground=self.colores["oro"],
                highlightthickness=1
            )
            tarjeta.grid(row=0, column=idx, padx=5, sticky="nsew")
            tarjetas.grid_columnconfigure(idx, weight=1)

            tk.Label(
                tarjeta,
                text=titulo,
                bg=self.colores["blanco_tarjeta"],
                fg="#777",
                font=("Arial", 9, "bold")
            ).pack(pady=(8, 2), padx=8)

            lbl_valor = tk.Label(
                tarjeta,
                text="0",
                bg=self.colores["blanco_tarjeta"],
                fg=self.colores["sidebar"],
                font=("Arial", 14, "bold")
            )
            lbl_valor.pack(pady=(0, 8), padx=8)
            valores_metricas[clave] = lbl_valor

        # =========================
        # BARRA DE ACCIONES
        # =========================
        acciones = tk.Frame(panel, bg=self.colores["fondo_main"])
        acciones.pack(fill="x", padx=25, pady=(0, 7))

        # =========================
        # TABLA
        # =========================
        tabla_frame = tk.Frame(panel, bg=self.colores["fondo_main"])
        tabla_frame.pack(fill="both", expand=True, padx=25, pady=(0, 10))

        columnas = (
            "ID", "Fecha", "Código", "Cliente", "Servicio", "Base USD",
            "IVA", "IGTF", "Total USD", "Tasa BCV", "Total Bs.",
            "Método", "Estado pago"
        )

        tabla = ttk.Treeview(
            tabla_frame,
            columns=columnas,
            show="headings",
            height=13
        )

        for col in columnas:
            tabla.heading(col, text=col)

            if col == "Cliente":
                tabla.column(col, width=160, anchor="w")
            elif col == "Servicio":
                tabla.column(col, width=190, anchor="w")
            elif col == "Método":
                tabla.column(col, width=105, anchor="center")
            elif col in ("Código", "Estado pago"):
                tabla.column(col, width=95, anchor="center")
            else:
                tabla.column(col, width=85, anchor="center")

        scroll_y = ttk.Scrollbar(tabla_frame, orient="vertical", command=tabla.yview)
        scroll_x = ttk.Scrollbar(tabla_frame, orient="horizontal", command=tabla.xview)
        tabla.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
        tabla.pack(fill="both", expand=True)

        # Guarda los datos crudos de cada fila cargada.
        datos_operaciones = {}

        def validar_fecha(valor):
            valor = (valor or "").strip()
            if not valor:
                return None

            try:
                datetime.datetime.strptime(valor, "%Y-%m-%d")
                return valor
            except ValueError:
                raise ValueError("Use el formato AAAA-MM-DD para las fechas.")

        def construir_where():
            condiciones = []
            params = []

            if combo_cliente.get() and combo_cliente.get() != "Todos":
                condiciones.append("c.nombre = ?")
                params.append(combo_cliente.get())

            if combo_servicio.get() and combo_servicio.get() != "Todos":
                condiciones.append("s.nombre_servicio = ?")
                params.append(combo_servicio.get())

            if combo_estado.get() and combo_estado.get() != "Todos":
                condiciones.append("COALESCE(ss.estado_pago,'Pendiente') = ?")
                params.append(combo_estado.get())

            desde = validar_fecha(entry_desde.get())
            hasta = validar_fecha(entry_hasta.get())

            if desde:
                condiciones.append("date(ss.fecha_solicitud) >= date(?)")
                params.append(desde)

            if hasta:
                condiciones.append("date(ss.fecha_solicitud) <= date(?)")
                params.append(hasta)

            where = ""
            if condiciones:
                where = " WHERE " + " AND ".join(condiciones)

            return where, params

        def cargar_filtros():
            try:
                conn = obtener_conexion()
                cur = conn.cursor()

                cur.execute(
                    """
                    SELECT DISTINCT nombre
                    FROM clientes
                    WHERE nombre IS NOT NULL
                      AND TRIM(nombre) != ''
                    ORDER BY nombre COLLATE NOCASE
                    """
                )
                clientes = [fila[0] for fila in cur.fetchall()]

                cur.execute(
                    """
                    SELECT nombre_servicio
                    FROM servicios
                    WHERE nombre_servicio IS NOT NULL
                      AND TRIM(nombre_servicio) != ''
                    ORDER BY nombre_servicio COLLATE NOCASE
                    """
                )
                servicios = [fila[0] for fila in cur.fetchall()]

                conn.close()

                combo_cliente["values"] = ["Todos"] + clientes
                combo_servicio["values"] = ["Todos"] + servicios

                combo_cliente.set("Todos")
                combo_servicio.set("Todos")

            except Exception as e:
                self._mensaje_corporativo(
                    "Filtros no disponibles",
                    "No se pudieron cargar los filtros de Gestión de Costos.",
                    tipo="error",
                    detalle=str(e)
                )

        def cargar_tabla():
            try:
                where, params = construir_where()
            except ValueError as e:
                self._mensaje_corporativo(
                    "Fecha inválida",
                    str(e),
                    tipo="advertencia"
                )
                return

            for item in tabla.get_children():
                tabla.delete(item)

            datos_operaciones.clear()

            try:
                conn = obtener_conexion()
                cur = conn.cursor()

                cur.execute(
                    f"""
                    SELECT
                        ss.id,
                        COALESCE(ss.fecha_solicitud,''),
                        COALESCE(c.codigo,''),
                        c.nombre,
                        s.nombre_servicio,
                        COALESCE(ss.monto,0),
                        COALESCE(ss.iva_usd,0),
                        COALESCE(ss.igtf_usd,0),
                        COALESCE(ss.total_usd,0),
                        COALESCE(ss.tasa_bcv,0),
                        COALESCE(ss.total_bs,0),
                        COALESCE(ss.metodo_pago,''),
                        COALESCE(ss.estado_pago,'Pendiente'),
                        COALESCE(ss.fecha_pago,''),
                        COALESCE(ss.observaciones,''),
                        c.correo,
                        c.telefono
                    FROM solicitudes_servicio ss
                    INNER JOIN clientes c ON c.id = ss.cliente_id
                    INNER JOIN servicios s ON s.id = ss.servicio_id
                    {where}
                    ORDER BY ss.id DESC
                    """,
                    params
                )

                filas = cur.fetchall()

                facturado = 0.0
                cobrado = 0.0
                por_cobrar = 0.0

                for f in filas:
                    operacion_id = int(f[0])
                    total = float(f[8] or 0)

                    facturado += total

                    if str(f[12]).lower() == "pagado":
                        cobrado += total
                    else:
                        por_cobrar += total

                    datos_operaciones[operacion_id] = {
                        "id": operacion_id,
                        "fecha": f[1],
                        "codigo": f[2],
                        "cliente": f[3],
                        "servicio": f[4],
                        "base": float(f[5] or 0),
                        "iva": float(f[6] or 0),
                        "igtf": float(f[7] or 0),
                        "total_usd": total,
                        "tasa_bcv": float(f[9] or 0),
                        "total_bs": float(f[10] or 0),
                        "metodo": f[11],
                        "estado": f[12],
                        "fecha_pago": f[13],
                        "observaciones": f[14],
                        "correo": f[15] or "",
                        "telefono": f[16] or "",
                    }

                    tabla.insert(
                        "",
                        "end",
                        values=(
                            f[0], f[1], f[2], f[3], f[4],
                            f"$ {float(f[5] or 0):,.2f}",
                            f"$ {float(f[6] or 0):,.2f}",
                            f"$ {float(f[7] or 0):,.2f}",
                            f"$ {total:,.2f}",
                            f"{float(f[9] or 0):.4f}",
                            f"Bs. {float(f[10] or 0):,.2f}",
                            f[11], f[12]
                        )
                    )

                conn.close()

                valores_metricas["operaciones"].config(text=str(len(filas)))
                valores_metricas["facturado"].config(text=f"$ {facturado:,.2f}")
                valores_metricas["cobrado"].config(text=f"$ {cobrado:,.2f}")
                valores_metricas["por_cobrar"].config(text=f"$ {por_cobrar:,.2f}")

            except Exception as e:
                self._mensaje_corporativo(
                    "Gestión de Costos no disponible",
                    "No se pudieron cargar las operaciones financieras.",
                    tipo="error",
                    detalle=str(e)
                )

        def limpiar_filtros():
            combo_cliente.set("Todos")
            combo_servicio.set("Todos")
            combo_estado.set("Todos")
            entry_desde.delete(0, "end")
            entry_hasta.delete(0, "end")
            cargar_tabla()

        def obtener_operacion_seleccionada():
            seleccion = tabla.selection()

            if not seleccion:
                self._mensaje_corporativo(
                    "Seleccione una operación",
                    "Primero seleccione una operación financiera de la tabla.",
                    tipo="advertencia"
                )
                return None

            valores = tabla.item(seleccion[0], "values")

            try:
                operacion_id = int(valores[0])
            except (ValueError, TypeError):
                return None

            return datos_operaciones.get(operacion_id)

        def ver_detalle():
            op = obtener_operacion_seleccionada()
            if not op:
                return

            modal, cuerpo = self._crear_modal_corporativo(
                f"Operación financiera #{op['id']}",
                "Detalle contable y estado de la cuenta por cobrar",
                ancho=680,
                alto=720,
                icono="$"
            )

            cuerpo.configure(
                bg=self.colores["blanco_tarjeta"],
                highlightbackground=self.colores["oro"],
                highlightthickness=1
            )

            campos = [
                ("Fecha solicitud", op["fecha"]),
                ("Código cliente", op["codigo"]),
                ("Cliente", op["cliente"]),
                ("Correo", op["correo"]),
                ("Teléfono", op["telefono"]),
                ("Servicio", op["servicio"]),
                ("Monto base", f"$ {op['base']:,.2f}"),
                ("IVA", f"$ {op['iva']:,.2f}"),
                ("IGTF", f"$ {op['igtf']:,.2f}"),
                ("Total USD", f"$ {op['total_usd']:,.2f}"),
                ("Tasa BCV", f"{op['tasa_bcv']:.4f}"),
                ("Total Bs.", f"Bs. {op['total_bs']:,.2f}"),
                ("Método de pago", op["metodo"]),
                ("Estado", op["estado"]),
                ("Fecha de pago", op["fecha_pago"] or "—"),
            ]

            for i, (etiqueta, valor) in enumerate(campos):
                tk.Label(
                    cuerpo,
                    text=f"{etiqueta}:",
                    bg=self.colores["blanco_tarjeta"],
                    fg=self.colores["sidebar"],
                    font=("Arial", 9, "bold")
                ).grid(row=i, column=0, sticky="nw", padx=(15, 12), pady=4)

                tk.Label(
                    cuerpo,
                    text=str(valor),
                    bg=self.colores["blanco_tarjeta"],
                    fg=self.colores["texto_oscuro"],
                    font=("Arial", 9),
                    wraplength=360,
                    justify="left"
                ).grid(row=i, column=1, sticky="nw", padx=(0, 15), pady=4)

            fila_obs = len(campos)

            tk.Label(
                cuerpo,
                text="Observaciones:",
                bg=self.colores["blanco_tarjeta"],
                fg=self.colores["sidebar"],
                font=("Arial", 9, "bold")
            ).grid(row=fila_obs, column=0, sticky="nw", padx=(15, 12), pady=(8, 4))

            txt_obs = tk.Text(
                cuerpo,
                width=45,
                height=5,
                wrap="word"
            )
            txt_obs.insert("1.0", op["observaciones"])
            txt_obs.grid(row=fila_obs, column=1, sticky="w", padx=(0, 15), pady=(8, 8))

            def guardar_observacion():
                observacion = txt_obs.get("1.0", "end").strip()

                try:
                    conn = obtener_conexion()
                    cur = conn.cursor()
                    cur.execute(
                        """
                        UPDATE solicitudes_servicio
                        SET observaciones = ?
                        WHERE id = ?
                        """,
                        (observacion, op["id"])
                    )
                    conn.commit()
                    conn.close()

                    registrar_accion(
                        getattr(self, "usuario_autenticado", "Sistema"),
                        "ACTUALIZÓ OBSERVACIÓN FINANCIERA",
                        f"Operación #{op['id']}."
                    )

                    self._mensaje_corporativo(
                        "Observación guardada",
                        "La observación fue actualizada correctamente.",
                        tipo="exito"
                    )

                    modal.destroy()
                    cargar_tabla()

                except Exception as e:
                    self._mensaje_corporativo(
                        "No se pudo guardar",
                        "Ocurrió un problema al actualizar la observación.",
                        tipo="error",
                        detalle=str(e)
                    )

            botones_detalle = tk.Frame(modal, bg=self.colores["fondo_main"])
            botones_detalle.pack(pady=(0, 18))

            tk.Button(
                botones_detalle,
                text="💾 Guardar observación",
                bg=self.colores["sidebar"],
                fg="white",
                font=("Arial", 10, "bold"),
                relief="flat",
                padx=14,
                pady=7,
                command=guardar_observacion
            ).pack(side="left", padx=5)

            tk.Button(
                botones_detalle,
                text="Cerrar",
                bg=self.colores["botones_menu"],
                fg=self.colores["oro"],
                font=("Arial", 10, "bold"),
                relief="flat",
                padx=14,
                pady=7,
                command=modal.destroy
            ).pack(side="left", padx=5)

        def cambiar_estado(nuevo_estado):
            op = obtener_operacion_seleccionada()
            if not op:
                return

            if op["estado"] == nuevo_estado:
                self._mensaje_corporativo(
                    "Sin cambios",
                    f"La operación ya está marcada como {nuevo_estado}.",
                    tipo="info"
                )
                return

            texto = (
                f"¿Desea marcar la operación #{op['id']} de "
                f"{op['cliente']} como {nuevo_estado}?"
            )

            if not self._confirmar_corporativo(
                "Cambiar estado de pago",
                texto,
                detalle=(
                    f"Cliente: {op['cliente']}\n"
                    f"Servicio: {op['servicio']}\n"
                    f"Total: $ {op['total_usd']:,.2f}"
                ),
                texto_confirmar=f"Marcar {nuevo_estado}",
                peligro=False
            ):
                return

            try:
                conn = obtener_conexion()
                cur = conn.cursor()

                if nuevo_estado == "Pagado":
                    cur.execute(
                        """
                        UPDATE solicitudes_servicio
                        SET estado_pago = 'Pagado',
                            fecha_pago = date('now')
                        WHERE id = ?
                        """,
                        (op["id"],)
                    )
                else:
                    cur.execute(
                        """
                        UPDATE solicitudes_servicio
                        SET estado_pago = 'Pendiente',
                            fecha_pago = NULL
                        WHERE id = ?
                        """,
                        (op["id"],)
                    )

                conn.commit()
                conn.close()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "ACTUALIZÓ ESTADO DE PAGO",
                    f"Operación #{op['id']}: {nuevo_estado}."
                )

                cargar_tabla()

            except Exception as e:
                self._mensaje_corporativo(
                    "No se pudo cambiar el estado",
                    "Ocurrió un problema al actualizar el estado de pago.",
                    tipo="error",
                    detalle=str(e)
                )

        def exportar_excel():
            if not datos_operaciones:
                self._mensaje_corporativo(
                    "Sin datos",
                    "No hay operaciones visibles para exportar.",
                    tipo="advertencia"
                )
                return

            ruta = filedialog.asksaveasfilename(
                title="Exportar Gestión de Costos a Excel",
                defaultextension=".xlsx",
                filetypes=[("Archivo Excel", "*.xlsx")]
            )

            if not ruta:
                return

            try:
                wb = Workbook()
                ws = wb.active
                assert ws is not None
                ws.title = "Gestión de Costos"

                encabezados = [
                    "ID", "Fecha", "Código cliente", "Cliente", "Servicio",
                    "Base USD", "IVA USD", "IGTF USD", "Total USD",
                    "Tasa BCV", "Total Bs.", "Método de pago",
                    "Estado", "Fecha pago", "Observaciones"
                ]
                ws.append(encabezados)

                for op in datos_operaciones.values():
                    ws.append([
                        op["id"], op["fecha"], op["codigo"], op["cliente"],
                        op["servicio"], op["base"], op["iva"], op["igtf"],
                        op["total_usd"], op["tasa_bcv"], op["total_bs"],
                        op["metodo"], op["estado"], op["fecha_pago"],
                        op["observaciones"]
                    ])

                for indice_columna, columna in enumerate(ws.columns, start=1):
                    ancho = 12
                    for celda in columna:
                        if celda.value is not None:
                            ancho = max(ancho, min(len(str(celda.value)) + 2, 45))
                    ws.column_dimensions[get_column_letter(indice_columna)].width = ancho

                wb.save(ruta)

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "EXPORTÓ GESTIÓN DE COSTOS",
                    f"Exportó {len(datos_operaciones)} operaciones a Excel."
                )

                self._mensaje_corporativo(
                    "Exportación completada",
                    "Gestión de Costos fue exportada correctamente a Excel.",
                    tipo="exito"
                )

            except Exception as e:
                self._mensaje_corporativo(
                    "No se pudo exportar",
                    "Ocurrió un problema al generar el archivo Excel.",
                    tipo="error",
                    detalle=str(e)
                )

        def exportar_pdf():
            if not datos_operaciones:
                self._mensaje_corporativo(
                    "Sin datos",
                    "No hay operaciones visibles para exportar.",
                    tipo="advertencia"
                )
                return

            ruta = filedialog.asksaveasfilename(
                title="Exportar Gestión de Costos a PDF",
                defaultextension=".pdf",
                filetypes=[("Archivo PDF", "*.pdf")]
            )

            if not ruta:
                return

            try:
                c = canvas.Canvas(ruta, pagesize=letter)
                ancho, alto = letter

                def encabezado_pdf():
                    c.setFont("Helvetica-Bold", 14)
                    c.drawString(40, alto - 45, "Macilitano Consulting Group C.A.")
                    c.setFont("Helvetica-Bold", 12)
                    c.drawString(40, alto - 65, "Reporte de Gestión de Costos")
                    c.setFont("Helvetica", 8)
                    c.drawString(
                        40, alto - 82,
                        f"Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                    return alto - 108

                y = encabezado_pdf()

                for op in datos_operaciones.values():
                    if y < 115:
                        c.showPage()
                        y = encabezado_pdf()

                    c.setFont("Helvetica-Bold", 9)
                    c.drawString(
                        40, y,
                        f"#{op['id']} | {op['cliente']} | {op['estado']}"
                    )
                    y -= 13

                    c.setFont("Helvetica", 8)
                    lineas = [
                        f"Fecha: {op['fecha']} | Código: {op['codigo']}",
                        f"Servicio: {op['servicio']}",
                        (
                            f"Base: ${op['base']:,.2f} | IVA: ${op['iva']:,.2f} | "
                            f"IGTF: ${op['igtf']:,.2f} | Total: ${op['total_usd']:,.2f}"
                        ),
                        (
                            f"BCV: {op['tasa_bcv']:.4f} | "
                            f"Total Bs.: {op['total_bs']:,.2f} | "
                            f"Método: {op['metodo']}"
                        ),
                    ]

                    for linea in lineas:
                        c.drawString(50, y, linea[:105])
                        y -= 11

                    y -= 7

                c.save()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "EXPORTÓ GESTIÓN DE COSTOS",
                    f"Exportó {len(datos_operaciones)} operaciones a PDF."
                )

                self._mensaje_corporativo(
                    "Exportación completada",
                    "Gestión de Costos fue exportada correctamente a PDF.",
                    tipo="exito"
                )

            except Exception as e:
                self._mensaje_corporativo(
                    "No se pudo exportar",
                    "Ocurrió un problema al generar el archivo PDF.",
                    tipo="error",
                    detalle=str(e)
                )

        # Acciones izquierda
        tk.Button(
            acciones,
            text="✓ Marcar Pagado",
            bg=self.colores["verde_aprobar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=11,
            pady=6,
            command=lambda: cambiar_estado("Pagado")
        ).pack(side="left")

        tk.Button(
            acciones,
            text="↩ Marcar Pendiente",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=11,
            pady=6,
            command=lambda: cambiar_estado("Pendiente")
        ).pack(side="left", padx=6)

        tk.Button(
            acciones,
            text="🔎 Ver detalle",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=11,
            pady=6,
            command=ver_detalle
        ).pack(side="left", padx=6)

        # Acciones derecha
        tk.Button(
            acciones,
            text="📄 PDF",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=11,
            pady=6,
            command=exportar_pdf
        ).pack(side="right")

        tk.Button(
            acciones,
            text="📗 Excel",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=11,
            pady=6,
            command=exportar_excel
        ).pack(side="right", padx=6)

        tk.Button(
            filtros,
            text="Limpiar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 9, "bold"),
            relief="flat",
            padx=8,
            command=limpiar_filtros
        ).pack(side="right", padx=3)

        tk.Button(
            filtros,
            text="Aplicar",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 9, "bold"),
            relief="flat",
            padx=8,
            command=cargar_tabla
        ).pack(side="right", padx=3)

        # Doble clic = detalle.
        tabla.bind("<Double-1>", lambda event: ver_detalle())

        cargar_filtros()
        cargar_tabla()

    # ----------------- GESTIÓN DE ASESORES -----------------
    def mostrar_asesores(self):

        if not self._requiere_administrador("administrar asesores y usuarios"):
            return
        panel = self.crear_panel_principal("Gestión de Asesores")

        top_bar = tk.Frame(panel, bg=self.colores["fondo_main"]) 
        top_bar.pack(fill="x", padx=25, pady=(5,10))

        btn_eliminar = tk.Button(top_bar, text="🗑 Eliminar", bg=self.colores["rojo_eliminar"], 
                            fg="white", font=("Arial", 10, "bold"), relief="flat", padx=10,
                                command=self.eliminar_asesor_bd)
        btn_eliminar.pack(side="right", padx=5)

        btn_editar = tk.Button(top_bar, text="✏ Editar", bg=self.colores["botones_menu"], 
                            fg=self.colores["oro"], font=("Arial", 10, "bold"), relief="flat", padx=10,
                            command=self.modal_editar_asesor)
        btn_editar.pack(side="right", padx=5)

        btn_agregar = tk.Button(top_bar, text="➕ Agregar Asesor", bg=self.colores["sidebar"], 
                                fg="white", font=("Arial", 10, "bold"), relief="flat", padx=10,
                                command=self.modal_agregar_asesor)
        btn_agregar.pack(side="right", padx=5)

        columnas = ("ID", "Usuario", "Nombre", "Correo", "Teléfono", "Especialidad", "Fecha Contratación", "Salario", "Activo")
        self.tabla_asesores = ttk.Treeview(panel, columns=columnas, show="headings", height=10)
        for col in columnas:
            self.tabla_asesores.heading(col, text=col)
            if col in ("Nombre", "Fecha Contratación"):
                self.tabla_asesores.column(col, anchor="w", width=150)
            elif col == "Salario":
                self.tabla_asesores.column(col, anchor="center", width=120)
            else:
                self.tabla_asesores.column(col, anchor="center", width=110)

        self.tabla_asesores.pack(fill="both", expand=True, padx=25, pady=(0,15))
        self.actualizar_tabla_asesores()

    def actualizar_tabla_asesores(self):
        if not hasattr(self, 'tabla_asesores'):
            return
        try:
            for item in self.tabla_asesores.get_children():
                self.tabla_asesores.delete(item)

            conn = obtener_conexion()
            cur = conn.cursor()
            cur.execute("SELECT a.id, u.username, a.nombre, a.correo, a.telefono, a.especialidad, a.fecha_contratacion, a.salario, a.activo FROM asesores a LEFT JOIN usuarios u ON a.usuario_id = u.id ORDER BY a.id")
            for fila in cur.fetchall():
                fila = list(fila)
                fila[7] = f"${fila[7]:,.2f}" if fila[7] is not None and fila[7] != 0 else ""
                fila[6] = fila[6] or ""
                self.tabla_asesores.insert("", "end", values=tuple(fila))
            conn.close()
        except Exception as e:
            self._mensaje_corporativo(
                "Asesores no disponibles",
                "No se pudo cargar la lista de asesores.",
                tipo="error",
                detalle=str(e)
            )

    def modal_agregar_asesor(self):

        if not self._requiere_administrador("crear asesores"):
            return

        modal, cuerpo = self._crear_modal_corporativo(
            "Nuevo asesor",
            "Crear perfil profesional y vincular acceso al sistema",
            ancho=690,
            alto=700,
            icono="＋",
        )

        form = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        form.pack(fill="both", expand=True)
        form.grid_columnconfigure(1, weight=1)

        def etiqueta(texto, fila):
            tk.Label(
                form,
                text=texto,
                bg=self.colores["fondo_main"],
                fg=self.colores["sidebar"],
                font=("Arial", 9, "bold"),
            ).grid(row=fila, column=0, sticky="w", padx=(0, 16), pady=7)

        def entrada(fila, ocultar=False):
            marco = tk.Frame(
                form,
                bg="white",
                highlightbackground="#d9d1c7",
                highlightthickness=1,
            )
            marco.grid(row=fila, column=1, sticky="ew", pady=6)
            e = tk.Entry(
                marco,
                font=("Arial", 10),
                relief="flat",
                bd=0,
                show="*" if ocultar else "",
            )
            e.pack(fill="x", padx=10, pady=8)
            return e

        etiqueta("Usuario de acceso", 0)
        ent_user = entrada(0)

        etiqueta("Contraseña", 1)
        ent_pw = entrada(1, ocultar=True)

        etiqueta("Nombre completo", 2)
        ent_nombre = entrada(2)

        etiqueta("Correo electrónico", 3)
        ent_correo = entrada(3)

        etiqueta("Teléfono", 4)
        ent_tel = entrada(4)

        etiqueta("Especialidad", 5)
        ent_esp = entrada(5)

        etiqueta("Fecha de contratación", 6)
        ent_fecha = entrada(6)
        ent_fecha.insert(0, datetime.date.today().strftime("%Y-%m-%d"))

        etiqueta("Salario mensual", 7)
        ent_salario = entrada(7)

        tk.Label(
            form,
            text=(
                "La cuenta creada tendrá rol Asesor. "
                "La fecha admite AAAA-MM-DD o DD/MM/AAAA."
            ),
            bg=self.colores["fondo_main"],
            fg="#766b6d",
            font=("Arial", 9),
            wraplength=540,
            justify="left",
        ).grid(
            row=8,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(10, 0),
        )

        def validar_fecha(texto):
            texto = texto.strip()
            if not texto:
                return False
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    datetime.datetime.strptime(texto, fmt)
                    return True
                except ValueError:
                    continue
            return False

        def validar_nombre_completo(texto):
            return len([p for p in texto.split() if p.strip()]) >= 2

        def validar_telefono(texto):
            digitos = re.sub(r"\D", "", texto)
            return len(digitos) == 11

        def validar_email(texto):
            return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", texto) is not None

        def aviso(titulo, mensaje):
            self._mensaje_corporativo(
                titulo,
                mensaje,
                tipo="advertencia",
            )

        def guardar():
            usuario = ent_user.get().strip()
            pw = ent_pw.get().strip()
            nombre = ent_nombre.get().strip()
            correo = ent_correo.get().strip()
            tel = ent_tel.get().strip()
            esp = ent_esp.get().strip()
            fecha_contratacion = ent_fecha.get().strip()
            salario_texto = ent_salario.get().strip()

            salario = 0.0
            if salario_texto:
                try:
                    salario = float(
                        salario_texto.replace("$", "").replace(",", "").strip()
                    )
                    if salario < 0:
                        raise ValueError
                except ValueError:
                    aviso("Salario inválido", "Ingrese un salario numérico válido.")
                    return

            if not usuario or not nombre:
                aviso(
                    "Datos incompletos",
                    "Usuario y nombre completo son obligatorios.",
                )
                return

            if not validar_nombre_completo(nombre):
                aviso(
                    "Nombre inválido",
                    "Ingrese al menos nombre y apellido.",
                )
                return

            if not validar_email(correo):
                aviso(
                    "Correo inválido",
                    "Ingrese una dirección de correo válida.",
                )
                return

            if not validar_telefono(tel):
                aviso(
                    "Teléfono inválido",
                    "Ingrese un número de teléfono con exactamente 11 dígitos.",
                )
                return

            if not validar_fecha(fecha_contratacion):
                aviso(
                    "Fecha inválida",
                    "Use el formato AAAA-MM-DD o DD/MM/AAAA.",
                )
                return

            try:
                conn = obtener_conexion()
                cur = conn.cursor()

                cur.execute(
                    "SELECT id FROM usuarios WHERE username = ?",
                    (usuario,),
                )
                row = cur.fetchone()

                if row:
                    usuario_id = int(row[0])
                else:
                    if not pw:
                        conn.close()
                        aviso(
                            "Contraseña requerida",
                            "Al crear un usuario nuevo debe indicar una contraseña.",
                        )
                        return

                    pw_store = hash_password(pw)
                    cur.execute(
                        """
                        INSERT INTO usuarios (username, password_hash, rol)
                        VALUES (?, ?, ?)
                        """,
                        (usuario, pw_store, "Asesor"),
                    )
                    ultimo_id = cur.lastrowid
                    if ultimo_id is None:
                        raise RuntimeError("No se pudo obtener el ID del usuario creado.")
                    usuario_id = int(ultimo_id)

                cur.execute(
                    """
                    SELECT id
                    FROM asesores
                    WHERE usuario_id = ?
                    """,
                    (usuario_id,),
                )
                if cur.fetchone():
                    conn.close()
                    aviso(
                        "Usuario ya vinculado",
                        "Ese usuario ya tiene un perfil de asesor asociado.",
                    )
                    return

                cur.execute(
                    """
                    INSERT INTO asesores (
                        usuario_id, nombre, correo, telefono,
                        especialidad, fecha_contratacion, salario
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        usuario_id,
                        nombre,
                        correo,
                        tel,
                        esp,
                        fecha_contratacion,
                        salario,
                    ),
                )

                conn.commit()
                conn.close()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "CREÓ ASESOR",
                    f"Creó el perfil de asesor de {nombre} ({usuario}).",
                )

                modal.destroy()
                self.actualizar_tabla_asesores()

                self._mensaje_corporativo(
                    "Asesor creado",
                    f"{nombre} fue incorporado correctamente.",
                    tipo="exito",
                    detalle=(
                        f"Usuario: {usuario}\n"
                        f"Especialidad: {esp or 'No especificada'}\n"
                        f"Fecha de contratación: {fecha_contratacion}"
                    ),
                )

            except Exception as e:
                detalle = traceback.format_exc()
                print(detalle)
                self._mensaje_corporativo(
                    "No se pudo crear el asesor",
                    "Ocurrió un problema al guardar el perfil.",
                    tipo="error",
                    detalle=str(e),
                )

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x", pady=(16, 0))

        tk.Button(
            botones,
            text="Cancelar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=modal.destroy,
        ).pack(side="right")

        tk.Button(
            botones,
            text="Guardar asesor",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=guardar,
        ).pack(side="right", padx=(0, 8))

        ent_user.focus_set()

    def modal_editar_asesor(self):

        if not self._requiere_administrador("editar asesores"):
            return

        sel = self.tabla_asesores.selection()
        if not sel:
            self._mensaje_corporativo(
                "Seleccione un asesor",
                "Primero seleccione un asesor de la tabla.",
                tipo="advertencia",
            )
            return

        vals = self.tabla_asesores.item(sel[0], "values")
        asesor_id = vals[0]

        try:
            conn = obtener_conexion()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT nombre, correo, telefono, especialidad,
                       fecha_contratacion, salario
                FROM asesores
                WHERE id = ?
                """,
                (asesor_id,),
            )
            row = cur.fetchone()
            conn.close()
        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo cargar el asesor",
                "Ocurrió un problema al consultar sus datos.",
                tipo="error",
                detalle=str(e),
            )
            return

        if not row:
            self._mensaje_corporativo(
                "Asesor no disponible",
                "El registro seleccionado ya no existe.",
                tipo="advertencia",
            )
            return

        modal, cuerpo = self._crear_modal_corporativo(
            "Editar asesor",
            f"Perfil interno #{asesor_id}",
            ancho=680,
            alto=610,
            icono="✎",
        )

        form = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        form.pack(fill="both", expand=True)
        form.grid_columnconfigure(1, weight=1)

        def etiqueta(texto, fila):
            tk.Label(
                form,
                text=texto,
                bg=self.colores["fondo_main"],
                fg=self.colores["sidebar"],
                font=("Arial", 9, "bold"),
            ).grid(row=fila, column=0, sticky="w", padx=(0, 16), pady=8)

        def entrada(fila, valor=""):
            marco = tk.Frame(
                form,
                bg="white",
                highlightbackground="#d9d1c7",
                highlightthickness=1,
            )
            marco.grid(row=fila, column=1, sticky="ew", pady=7)
            e = tk.Entry(marco, font=("Arial", 10), relief="flat", bd=0)
            e.pack(fill="x", padx=10, pady=8)
            if valor not in (None, ""):
                e.insert(0, str(valor))
            return e

        etiqueta("Nombre completo", 0)
        ent_nombre = entrada(0, row[0] or "")

        etiqueta("Correo electrónico", 1)
        ent_correo = entrada(1, row[1] or "")

        etiqueta("Teléfono", 2)
        ent_tel = entrada(2, row[2] or "")

        etiqueta("Especialidad", 3)
        ent_esp = entrada(3, row[3] or "")

        etiqueta("Fecha de contratación", 4)
        ent_fecha = entrada(4, row[4] or "")

        etiqueta("Salario mensual", 5)
        ent_salario = entrada(5, row[5] if row[5] is not None else "")

        tk.Label(
            form,
            text="La cuenta de acceso vinculada no se modifica desde esta ventana.",
            bg=self.colores["fondo_main"],
            fg="#766b6d",
            font=("Arial", 9),
        ).grid(
            row=6,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(10, 0),
        )

        def validar_fecha(texto):
            texto = texto.strip()
            if not texto:
                return False
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    datetime.datetime.strptime(texto, fmt)
                    return True
                except ValueError:
                    continue
            return False

        def validar_nombre_completo(texto):
            return len([p for p in texto.split() if p.strip()]) >= 2

        def validar_telefono(texto):
            return len(re.sub(r"\D", "", texto)) == 11

        def validar_email(texto):
            return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", texto) is not None

        def aviso(titulo, mensaje):
            self._mensaje_corporativo(
                titulo,
                mensaje,
                tipo="advertencia",
            )

        def guardar_edicion():
            nombre = ent_nombre.get().strip()
            correo = ent_correo.get().strip()
            tel = ent_tel.get().strip()
            esp = ent_esp.get().strip()
            fecha_contratacion = ent_fecha.get().strip()
            salario_texto = ent_salario.get().strip()

            try:
                salario = (
                    float(salario_texto.replace("$", "").replace(",", "").strip())
                    if salario_texto
                    else 0.0
                )
                if salario < 0:
                    raise ValueError
            except ValueError:
                aviso("Salario inválido", "Ingrese un salario numérico válido.")
                return

            if not nombre:
                aviso("Datos incompletos", "El nombre es obligatorio.")
                return

            if not validar_nombre_completo(nombre):
                aviso("Nombre inválido", "Ingrese al menos nombre y apellido.")
                return

            if not validar_email(correo):
                aviso("Correo inválido", "Ingrese una dirección de correo válida.")
                return

            if not validar_telefono(tel):
                aviso(
                    "Teléfono inválido",
                    "Ingrese un número de teléfono con exactamente 11 dígitos.",
                )
                return

            if not validar_fecha(fecha_contratacion):
                aviso(
                    "Fecha inválida",
                    "Use el formato AAAA-MM-DD o DD/MM/AAAA.",
                )
                return

            try:
                conn = obtener_conexion()
                cur = conn.cursor()
                cur.execute(
                    """
                    UPDATE asesores
                    SET nombre = ?, correo = ?, telefono = ?,
                        especialidad = ?, fecha_contratacion = ?, salario = ?
                    WHERE id = ?
                    """,
                    (
                        nombre,
                        correo,
                        tel,
                        esp,
                        fecha_contratacion,
                        salario,
                        asesor_id,
                    ),
                )
                conn.commit()
                conn.close()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "EDITÓ ASESOR",
                    f"Actualizó el perfil #{asesor_id} de {nombre}.",
                )

                modal.destroy()
                self.actualizar_tabla_asesores()

                self._mensaje_corporativo(
                    "Asesor actualizado",
                    f"Los datos de {nombre} fueron guardados correctamente.",
                    tipo="exito",
                )

            except Exception as e:
                detalle = traceback.format_exc()
                print(detalle)
                self._mensaje_corporativo(
                    "No se pudo actualizar",
                    "Ocurrió un problema al guardar los cambios.",
                    tipo="error",
                    detalle=str(e),
                )

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x", pady=(16, 0))

        tk.Button(
            botones,
            text="Cancelar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=modal.destroy,
        ).pack(side="right")

        tk.Button(
            botones,
            text="Guardar cambios",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=guardar_edicion,
        ).pack(side="right", padx=(0, 8))

        ent_nombre.focus_set()

    def eliminar_asesor_bd(self):

        if not self._requiere_administrador("eliminar asesores"):
            return

        sel = self.tabla_asesores.selection()
        if not sel:
            self._mensaje_corporativo(
                "Seleccione un asesor",
                "Primero seleccione el asesor que desea eliminar.",
                tipo="advertencia",
            )
            return

        vals = self.tabla_asesores.item(sel[0], "values")
        asesor_id = vals[0]
        usuario = vals[1] if len(vals) > 1 else "—"
        nombre = vals[2] if len(vals) > 2 else f"Asesor #{asesor_id}"

        if not self._confirmar_corporativo(
            "Eliminar asesor",
            f"¿Desea eliminar el perfil de {nombre}?",
            detalle=(
                f"ID interno: {asesor_id}\n"
                f"Usuario vinculado: {usuario}\n\n"
                "Se eliminará el perfil de asesor. La cuenta de usuario "
                "vinculada se conservará."
            ),
            texto_confirmar="Eliminar asesor",
            peligro=True,
        ):
            return

        try:
            conn = obtener_conexion()
            cur = conn.cursor()
            cur.execute(
                "DELETE FROM asesores WHERE id = ?",
                (asesor_id,),
            )
            conn.commit()
            conn.close()

            registrar_accion(
                getattr(self, "usuario_autenticado", "Sistema"),
                "ELIMINÓ ASESOR",
                f"Eliminó el perfil #{asesor_id} de {nombre}.",
            )

            self.actualizar_tabla_asesores()

            self._mensaje_corporativo(
                "Asesor eliminado",
                f"El perfil de {nombre} fue eliminado correctamente.",
                tipo="exito",
            )

        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo eliminar",
                "Ocurrió un problema al eliminar el asesor.",
                tipo="error",
                detalle=str(e),
            )

    def alternar_estado_servicio(self, event):
        """Alterna el estado nativo del cliente entre Activo e Inactivo con doble clic."""
        seleccion = self.tabla_clientes.selection()
        if not seleccion:
            return

        valores = self.tabla_clientes.item(seleccion[0], "values")
        id_cliente = valores[0]
        estado_actual = valores[7]  # Ajustado al nuevo índice en la tupla (Status es el índice 7)

        nuevo_estado = "Inactivo" if estado_actual == "Activo" else "Activo"

        try:
            conn = obtener_conexion()
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE clientes 
                SET estado = ? 
                WHERE id = ?
            """, (nuevo_estado, id_cliente))
            
            conn.commit()
            conn.close()
            
            self.actualizar_tabla_clientes()
        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo cambiar el estado",
                "Ocurrió un problema al actualizar el estado del cliente.",
                tipo="error",
                detalle=str(e)
            )

    def modal_agregar_cliente(self):
        """Registro manual de clientes con diseño corporativo."""
        modal, cuerpo = self._crear_modal_corporativo(
            "Nuevo cliente",
            "Registro manual en el directorio corporativo",
            ancho=650,
            alto=610,
            icono="＋",
        )

        form = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        form.pack(fill="both", expand=True)
        form.grid_columnconfigure(1, weight=1)

        def etiqueta(texto, fila):
            tk.Label(
                form,
                text=texto,
                bg=self.colores["fondo_main"],
                fg=self.colores["sidebar"],
                font=("Arial", 9, "bold"),
            ).grid(row=fila, column=0, sticky="w", padx=(0, 14), pady=8)

        def entrada(fila):
            marco = tk.Frame(
                form,
                bg="white",
                highlightbackground="#d9d1c7",
                highlightthickness=1,
            )
            marco.grid(row=fila, column=1, sticky="ew", pady=7)
            e = tk.Entry(marco, font=("Arial", 10), relief="flat", bd=0)
            e.pack(fill="x", padx=10, pady=8)
            return e

        # Catálogo real de servicios.
        servicios_disponibles = []
        try:
            conn = obtener_conexion()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT nombre_servicio
                FROM servicios
                WHERE nombre_servicio IS NOT NULL
                  AND TRIM(nombre_servicio) != ''
                ORDER BY nombre_servicio COLLATE NOCASE
                """
            )
            servicios_disponibles = [fila[0] for fila in cur.fetchall()]
            conn.close()
        except Exception:
            servicios_disponibles = []

        etiqueta("Cliente / empresa", 0)
        entry_nom = entrada(0)

        etiqueta("Sector / industria", 1)
        combo_ind = ttk.Combobox(
            form,
            state="readonly",
            values=TIPOS_EMPRESA,
            font=("Arial", 10),
        )
        combo_ind.grid(row=1, column=1, sticky="ew", pady=7, ipady=4)
        if TIPOS_EMPRESA:
            combo_ind.current(0)

        etiqueta("Servicio solicitado", 2)
        combo_serv = ttk.Combobox(
            form,
            state="readonly",
            values=servicios_disponibles,
            font=("Arial", 10),
        )
        combo_serv.grid(row=2, column=1, sticky="ew", pady=7, ipady=4)
        if servicios_disponibles:
            combo_serv.current(0)

        etiqueta("Teléfono", 3)
        entry_tel = entrada(3)

        etiqueta("Persona de contacto", 4)
        entry_per_con = entrada(4)

        etiqueta("Correo electrónico", 5)
        entry_cor = entrada(5)

        tk.Label(
            form,
            text="El código del cliente se generará automáticamente según su sector.",
            bg=self.colores["fondo_main"],
            fg="#766b6d",
            font=("Arial", 9),
            justify="left",
        ).grid(
            row=6,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(10, 4),
        )

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x", pady=(16, 0))

        def guardar_nuevo():
            nom = entry_nom.get().strip()
            cor = entry_cor.get().strip() or "No registrado"
            tel = entry_tel.get().strip() or "No registrado"
            per_con = entry_per_con.get().strip() or nom
            industria_sel = combo_ind.get().strip()
            servicio_sel = combo_serv.get().strip()

            if not nom:
                self._mensaje_corporativo(
                    "Falta información",
                    "El nombre del cliente o empresa es obligatorio.",
                    tipo="advertencia",
                )
                return

            if not industria_sel:
                self._mensaje_corporativo(
                    "Falta información",
                    "Seleccione el sector o industria del cliente.",
                    tipo="advertencia",
                )
                return

            if not servicio_sel:
                self._mensaje_corporativo(
                    "Falta información",
                    "Seleccione el servicio solicitado.",
                    tipo="advertencia",
                )
                return

            try:
                conn = obtener_conexion()
                cursor = conn.cursor()

                codigo = generar_id_cliente(industria_sel, conn)

                cursor.execute("PRAGMA table_info(clientes)")
                columnas = [col[1] for col in cursor.fetchall()]
                if "nombre_contacto" not in columnas:
                    cursor.execute(
                        "ALTER TABLE clientes ADD COLUMN nombre_contacto TEXT"
                    )
                    conn.commit()

                cursor.execute(
                    """
                    INSERT INTO clientes (
                        codigo, nombre, correo, telefono, industria,
                        servicio, nombre_contacto, fecha_registro, estado
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, date('now'), 'Activo')
                    """,
                    (
                        codigo,
                        nom,
                        cor,
                        tel,
                        industria_sel,
                        servicio_sel,
                        per_con,
                    ),
                )

                conn.commit()
                conn.close()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "AGREGÓ CLIENTE",
                    f"Registró a: {nom} [{industria_sel}] bajo el código {codigo}",
                )

                modal.destroy()
                self.actualizar_tabla_clientes()

                self._mensaje_corporativo(
                    "Cliente registrado",
                    f"{nom} fue agregado correctamente al directorio.",
                    tipo="exito",
                    detalle=f"Código asignado: {codigo}\nServicio: {servicio_sel}",
                )

            except Exception as e:
                self._mensaje_corporativo(
                    "No se pudo registrar el cliente",
                    "Ocurrió un problema al guardar la información.",
                    tipo="error",
                    detalle=str(e),
                )

        tk.Button(
            botones,
            text="Cancelar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=modal.destroy,
        ).pack(side="right")

        tk.Button(
            botones,
            text="Guardar cliente",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=guardar_nuevo,
        ).pack(side="right", padx=(0, 8))

        entry_nom.focus_set()

    def modal_editar_cliente(self):
        """Edición profesional de datos del cliente seleccionado."""
        seleccion = self.tabla_clientes.selection()

        if not seleccion:
            self._mensaje_corporativo(
                "Seleccione un cliente",
                "Primero seleccione un cliente de la tabla para editarlo.",
                tipo="advertencia",
            )
            return

        valores = self.tabla_clientes.item(seleccion[0], "values")
        id_cliente = valores[0]
        codigo_act = valores[1]
        nombre_act = valores[2]
        servicio_act = valores[3]
        telefono_act = valores[4]
        per_con_act = valores[5]
        correo_act = valores[6]

        industria_act = ""
        try:
            conn = obtener_conexion()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT industria FROM clientes WHERE id = ?",
                (id_cliente,),
            )
            fila = cursor.fetchone()
            if fila:
                industria_act = fila[0] or ""
            conn.close()
        except Exception as e:
            print(f"Error obteniendo detalles del cliente: {e}")

        servicios_disponibles = []
        try:
            conn = obtener_conexion()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT nombre_servicio
                FROM servicios
                WHERE nombre_servicio IS NOT NULL
                  AND TRIM(nombre_servicio) != ''
                ORDER BY nombre_servicio COLLATE NOCASE
                """
            )
            servicios_disponibles = [fila[0] for fila in cur.fetchall()]
            conn.close()
        except Exception:
            servicios_disponibles = []

        modal, cuerpo = self._crear_modal_corporativo(
            "Editar cliente",
            f"Código actual: {codigo_act}",
            ancho=650,
            alto=620,
            icono="✎",
        )

        form = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        form.pack(fill="both", expand=True)
        form.grid_columnconfigure(1, weight=1)

        def etiqueta(texto, fila):
            tk.Label(
                form,
                text=texto,
                bg=self.colores["fondo_main"],
                fg=self.colores["sidebar"],
                font=("Arial", 9, "bold"),
            ).grid(row=fila, column=0, sticky="w", padx=(0, 14), pady=8)

        def entrada(fila, valor=""):
            marco = tk.Frame(
                form,
                bg="white",
                highlightbackground="#d9d1c7",
                highlightthickness=1,
            )
            marco.grid(row=fila, column=1, sticky="ew", pady=7)
            e = tk.Entry(marco, font=("Arial", 10), relief="flat", bd=0)
            e.pack(fill="x", padx=10, pady=8)
            if valor not in (None, ""):
                e.insert(0, valor)
            return e

        etiqueta("Cliente / empresa", 0)
        entry_nom = entrada(0, nombre_act)

        etiqueta("Sector / industria", 1)
        combo_ind = ttk.Combobox(
            form,
            state="readonly",
            values=TIPOS_EMPRESA,
            font=("Arial", 10),
        )
        combo_ind.grid(row=1, column=1, sticky="ew", pady=7, ipady=4)

        if industria_act in TIPOS_EMPRESA:
            combo_ind.set(industria_act)
        elif TIPOS_EMPRESA:
            combo_ind.current(0)

        etiqueta("Servicio solicitado", 2)
        combo_serv = ttk.Combobox(
            form,
            state="readonly",
            values=servicios_disponibles,
            font=("Arial", 10),
        )
        combo_serv.grid(row=2, column=1, sticky="ew", pady=7, ipady=4)

        if servicio_act in servicios_disponibles:
            combo_serv.set(servicio_act)
        elif servicios_disponibles:
            combo_serv.current(0)

        etiqueta("Teléfono", 3)
        entry_tel = entrada(3, telefono_act)

        etiqueta("Persona de contacto", 4)
        entry_per_con = entrada(4, per_con_act)

        etiqueta("Correo electrónico", 5)
        entry_cor = entrada(5, correo_act)

        tk.Label(
            form,
            text=(
                "Si cambia el sector, el sistema generará un nuevo código "
                "acorde a la clasificación seleccionada."
            ),
            bg=self.colores["fondo_main"],
            fg="#766b6d",
            font=("Arial", 9),
            justify="left",
            wraplength=520,
        ).grid(
            row=6,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(10, 4),
        )

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x", pady=(16, 0))

        def actualizar_datos():
            nom = entry_nom.get().strip()
            cor = entry_cor.get().strip() or "No registrado"
            tel = entry_tel.get().strip() or "No registrado"
            per_con = entry_per_con.get().strip() or nom
            nueva_industria = combo_ind.get().strip()
            nuevo_servicio = combo_serv.get().strip()

            if not nom:
                self._mensaje_corporativo(
                    "Falta información",
                    "El nombre del cliente es obligatorio.",
                    tipo="advertencia",
                )
                return

            try:
                conn = obtener_conexion()
                cursor = conn.cursor()

                if nueva_industria != industria_act:
                    nuevo_codigo = generar_id_cliente(nueva_industria, conn)
                else:
                    nuevo_codigo = codigo_act

                cursor.execute(
                    """
                    UPDATE clientes
                    SET nombre = ?, correo = ?, telefono = ?, industria = ?,
                        servicio = ?, nombre_contacto = ?, codigo = ?
                    WHERE id = ?
                    """,
                    (
                        nom,
                        cor,
                        tel,
                        nueva_industria,
                        nuevo_servicio,
                        per_con,
                        nuevo_codigo,
                        id_cliente,
                    ),
                )

                conn.commit()
                conn.close()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "EDITÓ CLIENTE",
                    f"Modificó datos de {nom} (Código: {nuevo_codigo})",
                )

                modal.destroy()
                self.actualizar_tabla_clientes()

                self._mensaje_corporativo(
                    "Cliente actualizado",
                    f"Los datos de {nom} fueron guardados correctamente.",
                    tipo="exito",
                    detalle=f"Código final: {nuevo_codigo}\nServicio: {nuevo_servicio}",
                )

            except Exception as e:
                self._mensaje_corporativo(
                    "No se pudo actualizar",
                    "Ocurrió un problema al guardar los cambios.",
                    tipo="error",
                    detalle=str(e),
                )

        tk.Button(
            botones,
            text="Cancelar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=modal.destroy,
        ).pack(side="right")

        tk.Button(
            botones,
            text="Guardar cambios",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=actualizar_datos,
        ).pack(side="right", padx=(0, 8))

        entry_nom.focus_set()

    def eliminar_cliente_bd(self):

        if not self._requiere_administrador("eliminar clientes"):
            return

        seleccion = self.tabla_clientes.selection()
        if not seleccion:
            self._mensaje_corporativo(
                "Seleccione un cliente",
                "Primero seleccione un cliente del directorio.",
                tipo="advertencia",
            )
            return

        valores_raw = self.tabla_clientes.item(seleccion[0], "values")
        if not isinstance(valores_raw, (tuple, list)) or len(valores_raw) < 3:
            self._mensaje_corporativo(
                "Datos de cliente incompletos",
                "No fue posible leer correctamente el cliente seleccionado.",
                tipo="advertencia",
            )
            return

        valores = list(valores_raw)
        id_cliente = valores[0]
        codigo_cliente = valores[1] if len(valores) > 1 else "—"
        nombre_cliente = valores[2]

        if not self._confirmar_corporativo(
            "Eliminar cliente",
            f"¿Desea eliminar permanentemente a {nombre_cliente}?",
            detalle=(
                f"Código: {codigo_cliente}\n\n"
                "Esta acción eliminará también las operaciones financieras "
                "asociadas a este cliente y no puede deshacerse."
            ),
            texto_confirmar="Eliminar permanentemente",
            peligro=True,
        ):
            return

        try:
            conn = obtener_conexion()
            cursor = conn.cursor()

            cursor.execute(
                "DELETE FROM solicitudes_servicio WHERE cliente_id = ?",
                (id_cliente,),
            )
            cursor.execute(
                "DELETE FROM clientes WHERE id = ?",
                (id_cliente,),
            )

            conn.commit()
            conn.close()

            registrar_accion(
                getattr(self, "usuario_autenticado", "Sistema"),
                "ELIMINÓ CLIENTE",
                (
                    f"Removió de forma permanente a {nombre_cliente} "
                    f"(ID: {id_cliente}, código: {codigo_cliente})."
                ),
            )

            self.actualizar_tabla_clientes()

            self._mensaje_corporativo(
                "Cliente eliminado",
                f"{nombre_cliente} fue removido del sistema.",
                tipo="exito",
                detalle="También se eliminaron sus movimientos financieros asociados.",
            )

        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo eliminar",
                "Ocurrió un problema durante la eliminación del cliente.",
                tipo="error",
                detalle=str(e),
            )

    def mostrar_servicios(self):
        panel = self.crear_panel_principal("Catálogo de Servicios y Precios Base")

        tk.Label(
            panel,
            text="Los precios base se utilizan automáticamente al aprobar una asesoría web.",
            bg=self.colores["fondo_main"],
            fg=self.colores["texto_oscuro"],
            font=("Arial", 10, "italic")
        ).pack(anchor="w", padx=25, pady=(0, 8))

        top_bar = tk.Frame(panel, bg=self.colores["fondo_main"])
        top_bar.pack(fill="x", padx=25, pady=(0, 10))

        tabla_s: ttk.Treeview | None = None

        def servicio_seleccionado():
            if tabla_s is None:
                return None
            seleccion = tabla_s.selection()
            if not seleccion:
                self._mensaje_corporativo(
                    "Seleccione un servicio",
                    "Primero seleccione un servicio de la tabla.",
                    tipo="advertencia"
                )
                return None
            return tabla_s.item(seleccion[0], "values")

        def editar_servicio():
            if not self._requiere_administrador("editar servicios del catálogo"):
                return
            valores = servicio_seleccionado()
            if not valores:
                return

            servicio_id = int(valores[0])

            try:
                conn = obtener_conexion()
                cur = conn.cursor()
                cur.execute(
                    "SELECT nombre_servicio, descripcion, costo_base FROM servicios WHERE id = ?",
                    (servicio_id,)
                )
                fila = cur.fetchone()
                conn.close()
            except Exception as e:
                self._mensaje_corporativo(
                    "Servicio no disponible",
                    "No se pudo cargar la información del servicio.",
                    tipo="error",
                    detalle=str(e)
                )
                return

            if not fila:
                self._mensaje_corporativo(
                    "Servicio no disponible",
                    "El servicio seleccionado ya no existe.",
                    tipo="advertencia"
                )
                return

            modal, frame = self._crear_modal_corporativo(
                "Editar servicio",
                "Actualizar descripción y precio base",
                ancho=650,
                alto=520,
                icono="✎"
            )
            frame.grid_columnconfigure(1, weight=1)

            tk.Label(frame, text="Nombre:", bg=self.colores["fondo_main"], font=("Arial", 9, "bold")).grid(row=0, column=0, sticky="nw", pady=7)
            ent_nombre = tk.Entry(frame, width=48)
            ent_nombre.insert(0, fila[0] or "")
            ent_nombre.grid(row=0, column=1, sticky="ew", pady=7)

            tk.Label(frame, text="Descripción:", bg=self.colores["fondo_main"], font=("Arial", 9, "bold")).grid(row=1, column=0, sticky="nw", pady=7)
            txt_desc = tk.Text(frame, width=46, height=7, wrap="word")
            txt_desc.insert("1.0", fila[1] or "")
            txt_desc.grid(row=1, column=1, sticky="ew", pady=7)

            tk.Label(frame, text="Precio base (USD):", bg=self.colores["fondo_main"], font=("Arial", 9, "bold")).grid(row=2, column=0, sticky="w", pady=7)
            ent_costo = tk.Entry(frame, width=20)
            ent_costo.insert(0, f"{float(fila[2]):.2f}")
            ent_costo.grid(row=2, column=1, sticky="w", pady=7)

            def guardar_edicion():
                nombre = ent_nombre.get().strip()
                descripcion = txt_desc.get("1.0", "end").strip()
                costo_txt = ent_costo.get().strip().replace(",", ".")

                if not nombre or not descripcion or not costo_txt:
                    self._mensaje_corporativo("Datos incompletos", "Complete nombre, descripción y precio.", tipo="advertencia")
                    return

                try:
                    costo = float(costo_txt)
                    if costo <= 0:
                        raise ValueError
                except ValueError:
                    self._mensaje_corporativo("Precio inválido", "Ingrese un precio base mayor que cero.", tipo="advertencia")
                    return

                try:
                    conn = obtener_conexion()
                    cur = conn.cursor()
                    cur.execute(
                        """
                        SELECT id FROM servicios
                        WHERE LOWER(TRIM(nombre_servicio)) = LOWER(TRIM(?))
                          AND id <> ?
                        """,
                        (nombre, servicio_id)
                    )
                    if cur.fetchone():
                        conn.close()
                        self._mensaje_corporativo("Servicio duplicado", "Ya existe otro servicio con ese nombre.", tipo="advertencia")
                        return

                    cur.execute(
                        """
                        UPDATE servicios
                        SET nombre_servicio = ?, descripcion = ?, costo_base = ?
                        WHERE id = ?
                        """,
                        (nombre, descripcion, costo, servicio_id)
                    )
                    conn.commit()
                    conn.close()

                    registrar_accion(
                        getattr(self, "usuario_autenticado", "Sistema"),
                        "EDITÓ SERVICIO",
                        f"Actualizó '{nombre}' con precio base ${costo:,.2f}."
                    )

                    self._mensaje_corporativo("Servicio actualizado", "Los cambios fueron guardados correctamente.", tipo="exito")
                    modal.destroy()
                    self.mostrar_servicios()

                except Exception as e:
                    self._mensaje_corporativo("No se pudo actualizar", "Ocurrió un problema al guardar los cambios.", tipo="error", detalle=str(e))

            tk.Button(
                modal,
                text="Guardar cambios",
                bg=self.colores["verde_aprobar"],
                fg="white",
                font=("Arial", 10, "bold"),
                relief="flat",
                padx=18,
                pady=8,
                command=guardar_edicion
            ).pack(pady=16)

        def eliminar_servicio_seleccionado():
            if not self._requiere_administrador("eliminar servicios del catálogo"):
                return
            valores = servicio_seleccionado()
            if not valores:
                return

            servicio_id = int(valores[0])
            nombre = valores[1]

            if not self._confirmar_corporativo(
                "Eliminar servicio",
                f"¿Desea eliminar {nombre} del catálogo?",
                detalle=(
                    "La eliminación solo será posible si el servicio "
                    "no posee operaciones asociadas."
                ),
                texto_confirmar="Eliminar servicio",
                peligro=True
            ):
                return

            try:
                conn = obtener_conexion()
                cur = conn.cursor()

                cur.execute(
                    "SELECT COUNT(*) FROM solicitudes_servicio WHERE servicio_id = ?",
                    (servicio_id,)
                )
                usos = int(cur.fetchone()[0])

                if usos > 0:
                    conn.close()
                    self._mensaje_corporativo(
                        "Servicio en uso",
                        "Este servicio no puede eliminarse porque ya posee movimientos asociados.",
                        tipo="advertencia",
                        detalle="Puede editar su nombre, descripción o precio base sin perder el historial."
                    )
                    return

                cur.execute("DELETE FROM servicios WHERE id = ?", (servicio_id,))
                conn.commit()
                conn.close()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "ELIMINÓ SERVICIO",
                    f"Eliminó el servicio '{nombre}'."
                )

                self._mensaje_corporativo(
                    "Servicio eliminado",
                    f"{nombre} fue eliminado correctamente del catálogo.",
                    tipo="exito"
                )
                self.mostrar_servicios()

            except Exception as e:
                self._mensaje_corporativo(
                    "No se pudo eliminar",
                    "Ocurrió un problema al eliminar el servicio.",
                    tipo="error",
                    detalle=str(e)
                )

        btn_agregar_servicio = tk.Button(
            top_bar,
            text="➕ Agregar Servicio",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=12,
            pady=6,
            command=self.agregar_servicio
        )
        btn_agregar_servicio.pack(side="left")

        btn_editar_servicio = tk.Button(
            top_bar,
            text="✏ Editar Servicio",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=12,
            pady=6,
            command=editar_servicio
        )
        btn_editar_servicio.pack(side="left", padx=8)

        btn_eliminar_servicio = tk.Button(
            top_bar,
            text="🗑 Eliminar Servicio",
            bg=self.colores["rojo_eliminar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=12,
            pady=6,
            command=eliminar_servicio_seleccionado
        )
        btn_eliminar_servicio.pack(side="left")

        if not self._es_administrador():
            btn_agregar_servicio.pack_forget()
            btn_editar_servicio.pack_forget()
            btn_eliminar_servicio.pack_forget()

        columnas = ("ID", "Nombre Servicio", "Descripción", "Precio Base USD")
        tabla_s = ttk.Treeview(panel, columns=columnas, show="headings", height=15)

        tabla_s.heading("ID", text="ID")
        tabla_s.heading("Nombre Servicio", text="Servicio")
        tabla_s.heading("Descripción", text="Descripción")
        tabla_s.heading("Precio Base USD", text="Precio Base")

        tabla_s.column("ID", width=55, anchor="center", stretch=False)
        tabla_s.column("Nombre Servicio", width=270, anchor="w")
        tabla_s.column("Descripción", width=520, anchor="w")
        tabla_s.column("Precio Base USD", width=130, anchor="e", stretch=False)

        scroll_y = ttk.Scrollbar(panel, orient="vertical", command=tabla_s.yview)
        tabla_s.configure(yscrollcommand=scroll_y.set)

        scroll_y.pack(side="right", fill="y", pady=10, padx=(0, 20))
        tabla_s.pack(fill="both", expand=True, padx=(25, 0), pady=10)

        try:
            conn = obtener_conexion()
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT id, nombre_servicio, descripcion, costo_base
                FROM servicios
                ORDER BY nombre_servicio COLLATE NOCASE
                """
            )
            for fila in cursor.fetchall():
                tabla_s.insert(
                    "",
                    "end",
                    values=(
                        fila[0],
                        fila[1],
                        fila[2],
                        f"$ {float(fila[3]):,.2f}"
                    )
                )
            conn.close()
        except Exception as e:
            self._mensaje_corporativo(
                "Catálogo no disponible",
                "No se pudo cargar el catálogo de servicios.",
                tipo="error",
                detalle=str(e)
            )

    def agregar_servicio(self):
        if not self._requiere_administrador("agregar servicios al catálogo"):
            return

        modal, cuerpo = self._crear_modal_corporativo(
            "Nuevo servicio",
            "Agregar una opción al catálogo corporativo",
            ancho=650,
            alto=520,
            icono="＋",
        )

        form = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        form.pack(fill="both", expand=True)
        form.grid_columnconfigure(1, weight=1)

        tk.Label(
            form,
            text="Nombre del servicio",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 9, "bold"),
        ).grid(row=0, column=0, sticky="nw", padx=(0, 14), pady=8)

        marco_nombre = tk.Frame(
            form,
            bg="white",
            highlightbackground="#d9d1c7",
            highlightthickness=1,
        )
        marco_nombre.grid(row=0, column=1, sticky="ew", pady=7)

        entry_nombre = tk.Entry(
            marco_nombre,
            font=("Arial", 10),
            relief="flat",
            bd=0,
        )
        entry_nombre.pack(fill="x", padx=10, pady=8)

        tk.Label(
            form,
            text="Descripción",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 9, "bold"),
        ).grid(row=1, column=0, sticky="nw", padx=(0, 14), pady=8)

        txt_desc = tk.Text(
            form,
            height=7,
            wrap="word",
            font=("Arial", 10),
            relief="solid",
            bd=1,
        )
        txt_desc.grid(row=1, column=1, sticky="ew", pady=7)

        tk.Label(
            form,
            text="Precio base (USD)",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 9, "bold"),
        ).grid(row=2, column=0, sticky="w", padx=(0, 14), pady=8)

        marco_precio = tk.Frame(
            form,
            bg="white",
            highlightbackground="#d9d1c7",
            highlightthickness=1,
        )
        marco_precio.grid(row=2, column=1, sticky="ew", pady=7)

        entry_costo = tk.Entry(
            marco_precio,
            font=("Arial", 10),
            relief="flat",
            bd=0,
        )
        entry_costo.pack(fill="x", padx=10, pady=8)

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x", pady=(18, 0))

        def guardar():
            nombre = entry_nombre.get().strip()
            desc = txt_desc.get("1.0", "end").strip()
            costo_txt = entry_costo.get().strip().replace(",", ".")

            if not nombre or not desc or not costo_txt:
                self._mensaje_corporativo(
                    "Datos incompletos",
                    "Complete nombre, descripción y precio base.",
                    tipo="advertencia",
                )
                return

            try:
                costo = float(costo_txt)
                if costo <= 0:
                    raise ValueError
            except ValueError:
                self._mensaje_corporativo(
                    "Precio inválido",
                    "Ingrese un precio base mayor que cero.",
                    tipo="advertencia",
                )
                return

            try:
                conn = obtener_conexion()
                cursor = conn.cursor()

                cursor.execute(
                    """
                    SELECT id FROM servicios
                    WHERE LOWER(TRIM(nombre_servicio)) = LOWER(TRIM(?))
                    """,
                    (nombre,),
                )

                if cursor.fetchone():
                    conn.close()
                    self._mensaje_corporativo(
                        "Servicio duplicado",
                        "Ya existe un servicio con ese nombre.",
                        tipo="advertencia",
                    )
                    return

                cursor.execute(
                    """
                    INSERT INTO servicios (
                        nombre_servicio, descripcion, costo_base
                    )
                    VALUES (?, ?, ?)
                    """,
                    (nombre, desc, costo),
                )

                conn.commit()
                conn.close()

                registrar_accion(
                    getattr(self, "usuario_autenticado", "Sistema"),
                    "AGREGÓ SERVICIO",
                    f"Agregó '{nombre}' con precio base ${costo:,.2f}.",
                )

                modal.destroy()
                self.mostrar_servicios()

                self._mensaje_corporativo(
                    "Servicio agregado",
                    f"{nombre} fue incorporado al catálogo.",
                    tipo="exito",
                    detalle=f"Precio base: $ {costo:,.2f}",
                )

            except Exception as e:
                self._mensaje_corporativo(
                    "No se pudo agregar el servicio",
                    "Ocurrió un problema al guardar la información.",
                    tipo="error",
                    detalle=str(e),
                )

        tk.Button(
            botones,
            text="Cancelar",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=modal.destroy,
        ).pack(side="right")

        tk.Button(
            botones,
            text="Guardar servicio",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            command=guardar,
        ).pack(side="right", padx=(0, 8))

        entry_nombre.focus_set()

    def mostrar_notificaciones(self):
        panel = self.crear_panel_principal("Solicitudes y Citas Web")

        tk.Label(
            panel,
            text="Bandeja activa de formularios recibidos desde la página web.",
            bg=self.colores["fondo_main"],
            font=("Arial", 10, "italic")
        ).pack(anchor="w", padx=25, pady=(0, 8))

        self._mostrar_alerta_servicios_nuevos(panel)

        barra = tk.Frame(panel, bg=self.colores["fondo_main"])
        barra.pack(fill="x", padx=25, pady=(0, 8))

        tk.Button(
            barra,
            text="🕘 Historial de Solicitudes",
            bg=self.colores["botones_menu"],
            fg=self.colores["oro"],
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=12,
            pady=7,
            command=self.mostrar_historial_solicitudes
        ).pack(side="right")

        columnas_w = (
            "ID",
            "Cliente Potencial",
            "Servicio de Interés",
            "Descripción del Negocio",
            "Estado",
            "Correo",
            "Teléfono"
        )

        tabla_web = ttk.Treeview(
            panel,
            columns=columnas_w,
            show="headings",
            height=12
        )

        for col in columnas_w:
            tabla_web.heading(col, text=col)

            if col == "Descripción del Negocio":
                tabla_web.column(col, anchor="w", width=280)
            elif col in ("Correo", "Teléfono"):
                tabla_web.column(col, anchor="w", width=160)
            elif col == "ID":
                tabla_web.column(col, anchor="center", width=85)
            else:
                tabla_web.column(col, anchor="center", width=145)

        tabla_web.pack(fill="both", expand=True, padx=25, pady=10)

        frame_botones = tk.Frame(panel, bg=self.colores["fondo_main"])
        frame_botones.pack(fill="x", padx=25, pady=10)

        btn_aprobar = tk.Button(
            frame_botones,
            text="✔ Aprobar Asesoría",
            bg=self.colores["verde_aprobar"],
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=15,
            pady=8,
            command=lambda: self.aprobar_y_calcular_solicitud(tabla_web)
        )
        btn_aprobar.pack(side="left")

        btn_rechazar = tk.Button(
            frame_botones,
            text="✖ Rechazar Servicio",
            bg="#9e2a2b",
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=15,
            pady=8,
            command=lambda: self.rechazar_solicitud(tabla_web)
        )
        btn_rechazar.pack(side="left", padx=8)

        try:
            data = self._api_solicitudes("GET", "/api/solicitudes")

            for fila in data.get("items", []):
                estado = fila.get("estado") or "Pendiente"

                # Solicitudes aún no finalizadas; se conserva compatibilidad con registros antiguos en Contactado.
                if estado.lower() not in {"pendiente", "contactado"}:
                    continue

                tabla_web.insert(
                    "",
                    "end",
                    values=(
                        f"ID Web: {fila.get('id')}",
                        fila.get("cliente_potencial") or "No registrado",
                        fila.get("servicio_interes") or "No especificado",
                        fila.get("descripcion") or "",
                        estado,
                        fila.get("correo") or "No registrado",
                        fila.get("telefono") or "No registrado"
                    )
                )

        except Exception as e:
            tabla_web.insert(
                "",
                "end",
                values=("Error", "Servidor web", str(e), "", "", "", "")
            )

    def mostrar_historial_solicitudes(self):
        panel = self.crear_panel_principal("Historial de Solicitudes Web")

        tk.Label(
            panel,
            text="Consulta de solicitudes procesadas y su trazabilidad.",
            bg=self.colores["fondo_main"],
            font=("Arial", 10, "italic")
        ).pack(anchor="w", padx=25, pady=(0, 10))

        filtros = tk.Frame(panel, bg=self.colores["fondo_main"])
        filtros.pack(fill="x", padx=25, pady=(0, 10))

        tk.Label(
            filtros,
            text="Estado:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 10, "bold")
        ).pack(side="left")

        combo_estado = ttk.Combobox(
            filtros,
            state="readonly",
            width=18,
            values=["Todos", "Pendiente", "Contactado", "Aprobado", "Rechazado"]
        )
        combo_estado.set("Todos")
        combo_estado.pack(side="left", padx=(6, 15))

        tk.Label(
            filtros,
            text="Buscar:",
            bg=self.colores["fondo_main"],
            fg=self.colores["sidebar"],
            font=("Arial", 10, "bold")
        ).pack(side="left")

        entrada_buscar = tk.Entry(filtros, width=28)
        entrada_buscar.pack(side="left", padx=6, ipady=3)

        columnas = (
            "ID",
            "Fecha solicitud",
            "Cliente",
            "Servicio",
            "Estado",
            "Procesado por",
            "Última actualización",
            "Correo",
            "Teléfono"
        )

        tabla = ttk.Treeview(
            panel,
            columns=columnas,
            show="headings",
            height=16
        )

        for col in columnas:
            tabla.heading(col, text=col)

            if col in ("Cliente", "Servicio", "Correo"):
                tabla.column(col, width=180, anchor="w")
            elif col == "ID":
                tabla.column(col, width=60, anchor="center")
            else:
                tabla.column(col, width=125, anchor="center")

        sb = ttk.Scrollbar(panel, orient="vertical", command=tabla.yview)
        tabla.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y", padx=(0, 18), pady=5)
        tabla.pack(fill="both", expand=True, padx=(25, 0), pady=5)

        def cargar():
            for item in tabla.get_children():
                tabla.delete(item)

            try:
                data = self._api_solicitudes("GET", "/api/solicitudes")
                estado_filtro = combo_estado.get().strip()
                texto = entrada_buscar.get().strip().lower()

                for fila in data.get("items", []):
                    estado = fila.get("estado") or "Pendiente"

                    if estado_filtro != "Todos" and estado != estado_filtro:
                        continue

                    cadena = " ".join([
                        str(fila.get("cliente_potencial") or ""),
                        str(fila.get("servicio_interes") or ""),
                        str(fila.get("correo") or ""),
                        str(fila.get("telefono") or ""),
                        str(fila.get("procesado_por") or "")
                    ]).lower()

                    if texto and texto not in cadena:
                        continue

                    tabla.insert(
                        "",
                        "end",
                        values=(
                            fila.get("id"),
                            fila.get("fecha_solicitud") or "",
                            fila.get("cliente_potencial") or "",
                            fila.get("servicio_interes") or "",
                            estado,
                            fila.get("procesado_por") or "—",
                            fila.get("actualizado_en") or "—",
                            fila.get("correo") or "",
                            fila.get("telefono") or ""
                        )
                    )

            except Exception as e:
                self._mensaje_corporativo(
                    "Historial no disponible",
                    "No se pudo cargar el historial de solicitudes.",
                    tipo="error",
                    detalle=str(e)
                )

        combo_estado.bind("<<ComboboxSelected>>", lambda e: cargar())
        entrada_buscar.bind("<KeyRelease>", lambda e: cargar())

        tk.Button(
            filtros,
            text="↻ Actualizar",
            bg=self.colores["sidebar"],
            fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=10,
            command=cargar
        ).pack(side="right")

        cargar()

    def rechazar_solicitud(self, tabla):
        item_seleccionado = tabla.selection()

        if not item_seleccionado:
            self._mensaje_corporativo(
                "Seleccione una solicitud",
                "Primero seleccione la solicitud que desea rechazar.",
                tipo="advertencia",
            )
            return

        valores = tabla.item(item_seleccionado, "values")
        estado_actual = valores[4]

        if str(estado_actual).lower() not in {"pendiente", "contactado"}:
            self._mensaje_corporativo(
                "Solicitud procesada",
                "Esta solicitud ya fue procesada anteriormente.",
                tipo="info",
            )
            return

        ref_id = valores[0]
        cliente = valores[1] if len(valores) > 1 else "Cliente"
        servicio = valores[2] if len(valores) > 2 else "—"
        id_web_real = ref_id.replace("ID Web: ", "")

        if not self._confirmar_corporativo(
            "Rechazar solicitud",
            f"¿Desea rechazar la solicitud de {cliente}?",
            detalle=(
                f"Servicio solicitado: {servicio}\n"
                f"Referencia: {ref_id}\n\n"
                "La solicitud pasará al historial como Rechazada."
            ),
            texto_confirmar="Rechazar solicitud",
            peligro=True,
        ):
            return

        try:
            self._api_solicitudes(
                "PATCH",
                f"/api/solicitudes/{int(id_web_real)}",
                json={"estado": "Rechazado"},
            )

            registrar_accion(
                getattr(self, "usuario_autenticado", "Sistema"),
                "RECHAZÓ SOLICITUD WEB",
                f"Rechazó la solicitud web ID {id_web_real}.",
            )

            self.mostrar_notificaciones()

            self._mensaje_corporativo(
                "Solicitud rechazada",
                "La solicitud fue movida correctamente al historial.",
                tipo="exito",
                detalle=f"Cliente: {cliente}\nServicio: {servicio}",
            )

        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo rechazar",
                "Ocurrió un problema al procesar la solicitud.",
                tipo="error",
                detalle=str(e),
            )

    def _mostrar_alerta_servicios_nuevos(self, panel):
        try:
            conn = obtener_conexion()
            cursor = conn.cursor()
            cursor.execute("SELECT id, nombre_servicio FROM servicios ORDER BY id DESC LIMIT 5")
            servicios = cursor.fetchall()
            conn.close()
        except Exception:
            servicios = []

        if not servicios:
            return

        frame_alerta = tk.Frame(panel, bg="#fff3cd", highlightbackground="#f0b429", highlightthickness=1)
        frame_alerta.pack(fill="x", padx=25, pady=(0, 10))

        tk.Label(frame_alerta, text="📣 Nuevos servicios en el catálogo", bg="#fff3cd", fg="#7a4b00", font=("Arial", 11, "bold")).pack(anchor="w", padx=10, pady=(8, 2))
        for servicio_id, nombre in servicios:
            tk.Label(frame_alerta, text=f"• {nombre}", bg="#fff3cd", fg="#3b2f2f", font=("Arial", 10)).pack(anchor="w", padx=14, pady=1)

    def seleccionar_metodo_pago(self):
        ventana, cuerpo = self._crear_modal_corporativo(
            "Método de pago",
            "Información financiera de la asesoría",
            ancho=560,
            alto=330,
            icono="$"
        )

        tk.Label(
            cuerpo, text="Seleccione cómo realizará el pago el cliente.",
            bg=self.colores["fondo_main"], fg=self.colores["texto_oscuro"],
            font=("Arial", 11, "bold")
        ).pack(anchor="w", pady=(4, 16))

        resultado: dict[str, str | None] = {"valor": None}

        def elegir(valor):
            resultado["valor"] = valor
            ventana.destroy()

        opciones = [
            ("Transferencia", "transferencia", "Transferencia bancaria"),
            ("Efectivo", "efectivo", "Pago realizado en efectivo"),
            ("Pago Móvil", "pago movil", "Pago móvil / transferencia inmediata"),
        ]

        for titulo, valor, descripcion in opciones:
            tk.Button(
                cuerpo,
                text=f"{titulo}\n{descripcion}",
                bg="white",
                fg=self.colores["sidebar"],
                activebackground="#f7efe9",
                activeforeground=self.colores["sidebar"],
                font=("Arial", 10, "bold"),
                relief="flat",
                anchor="w",
                justify="left",
                padx=14,
                pady=9,
                cursor="hand2",
                command=lambda v=valor: elegir(v)
            ).pack(fill="x", pady=3)

        tk.Button(
            cuerpo, text="Cancelar", bg=self.colores["botones_menu"],
            fg=self.colores["oro"], font=("Arial", 10, "bold"),
            relief="flat", padx=18, pady=7, command=ventana.destroy
        ).pack(anchor="e", pady=(12, 0))

        ventana.wait_window()
        return resultado["valor"]

    def seleccionar_industria_cliente(self):
        ventana, cuerpo = self._crear_modal_corporativo(
            "Clasificación de empresa",
            "Datos comerciales del nuevo cliente",
            ancho=590,
            alto=340,
            icono="▦"
        )

        tk.Label(
            cuerpo, text="Seleccione la actividad o sector principal de la empresa.",
            bg=self.colores["fondo_main"], fg=self.colores["texto_oscuro"],
            font=("Arial", 11, "bold")
        ).pack(anchor="w", pady=(4, 6))

        tk.Label(
            cuerpo,
            text="Esta clasificación se utiliza para generar automáticamente el código del cliente.",
            bg=self.colores["fondo_main"], fg="#766b6d",
            font=("Arial", 9), wraplength=520, justify="left"
        ).pack(anchor="w", pady=(0, 14))

        combo = ttk.Combobox(
            cuerpo, state="readonly", values=TIPOS_EMPRESA,
            font=("Arial", 10), width=48
        )
        combo.pack(fill="x", ipady=5)
        combo.current(0)

        resultado: dict[str, str | None] = {"valor": None}

        def confirmar(event=None):
            valor = combo.get().strip()
            if not valor:
                return
            resultado["valor"] = valor
            ventana.destroy()

        botones = tk.Frame(cuerpo, bg=self.colores["fondo_main"])
        botones.pack(fill="x", pady=(22, 0))

        tk.Button(
            botones, text="Cancelar", bg=self.colores["botones_menu"],
            fg=self.colores["oro"], font=("Arial", 10, "bold"),
            relief="flat", padx=18, pady=7, command=ventana.destroy
        ).pack(side="right")

        tk.Button(
            botones, text="Confirmar clasificación", bg=self.colores["sidebar"],
            fg="white", font=("Arial", 10, "bold"), relief="flat",
            padx=18, pady=7, command=confirmar
        ).pack(side="right", padx=(0, 8))

        combo.bind("<Return>", confirmar)
        ventana.wait_window()
        return resultado["valor"]

    def aprobar_y_calcular_solicitud(self, tabla):
        item_seleccionado = tabla.selection()
        if not item_seleccionado:
            self._mensaje_corporativo(
                "Seleccione una solicitud",
                "Primero seleccione una solicitud de la lista.",
                tipo="advertencia"
            )
            return
            
        valores = tabla.item(item_seleccionado, "values")
        ref_id, cliente_nombre, servicio_interes, descripcion_negocio, estado_actual, correo_real, telefono_real = valores

        if str(estado_actual).lower() not in {"pendiente", "contactado"}:
            self._mensaje_corporativo(
                "Solicitud procesada",
                "Esta solicitud ya fue procesada anteriormente.",
                tipo="info"
            )
            return

        id_web_real = ref_id.replace("ID Web: ", "")

        industria_cliente = self.seleccionar_industria_cliente()
        if not industria_cliente:
            return

        tasa_bcv = obtener_tasa_bcv_automatica()
        if not tasa_bcv:
            tasa_bcv_input = self._dialogo_texto_corporativo(
                "Tasa BCV manual",
                "No fue posible obtener la tasa automáticamente.",
                etiqueta="Ingrese la tasa BCV (Bs/USD):"
            )
            if not tasa_bcv_input:
                return
            try:
                tasa_bcv = float(tasa_bcv_input.replace(",", "."))
                if tasa_bcv <= 0:
                    raise ValueError
            except ValueError:
                self._mensaje_corporativo(
                    "Tasa inválida",
                    "La tasa BCV debe ser un número mayor que cero.",
                    tipo="advertencia"
                )
                return

        metodo_pago = self.seleccionar_metodo_pago()
        if not metodo_pago:
            return

        try:
            conn_local = obtener_conexion()
            cursor_local = conn_local.cursor()

            codigo_cliente = generar_id_cliente(industria_cliente, conn_local)
            nombre_contacto = self._dialogo_texto_corporativo(
                "Persona de contacto",
                "Indique quién será la persona de contacto principal de este cliente.",
                etiqueta="Nombre y apellido:",
                valor_inicial=cliente_nombre
            )
            if not nombre_contacto:
                conn_local.close()
                return

            cursor_local.execute("""
                INSERT INTO clientes (codigo, nombre, correo, telefono, industria, servicio, nombre_contacto, fecha_registro, estado)
                VALUES (?, ?, ?, ?, ?, ?, ?, date('now'), 'Activo')
            """, (codigo_cliente, cliente_nombre, correo_real, telefono_real, industria_cliente, servicio_interes, nombre_contacto))
            
            cliente_id = cursor_local.lastrowid

            # Resolver el servicio exacto del catálogo.
            # Se mantiene compatibilidad con el antiguo texto "Sistema Holgado".
            aliases_servicio = {
                "Transformación Digital (Sistema Holgado)": "Transformación Digital (Sistema Homologado)",
            }
            servicio_catalogo = aliases_servicio.get(servicio_interes, servicio_interes)

            cursor_local.execute(
                """
                SELECT id, costo_base
                FROM servicios
                WHERE LOWER(TRIM(nombre_servicio)) = LOWER(TRIM(?))
                """,
                (servicio_catalogo,)
            )
            res_serv = cursor_local.fetchone()

            # Si llegó "Otro servicio" desde la web, el asesor define el precio
            # una sola vez y el servicio queda incorporado al catálogo.
            if not res_serv:
                precio_personalizado = self._dialogo_texto_corporativo(
                    "Nuevo servicio detectado",
                    (
                        "Este servicio todavía no existe en el catálogo.\n\n"
                        f"{servicio_interes}"
                    ),
                    etiqueta="Precio base en USD:"
                )
                if not precio_personalizado:
                    conn_local.close()
                    return

                try:
                    monto_base_usd = float(
                        precio_personalizado.strip().replace(",", ".")
                    )
                    if monto_base_usd <= 0:
                        raise ValueError
                except ValueError:
                    conn_local.close()
                    self._mensaje_corporativo(
                        "Precio inválido",
                        "El precio base debe ser un número mayor que cero.",
                        tipo="advertencia"
                    )
                    return

                cursor_local.execute(
                    """
                    INSERT INTO servicios (nombre_servicio, descripcion, costo_base)
                    VALUES (?, ?, ?)
                    """,
                    (
                        servicio_interes,
                        "Servicio incorporado desde una solicitud personalizada de la página web.",
                        monto_base_usd
                    )
                )
                servicio_id = cursor_local.lastrowid

            else:
                servicio_id = int(res_serv[0])
                monto_base_usd = float(res_serv[1])

            config_fin = self._obtener_configuracion_financiera()
            iva_tasa = config_fin["iva_pct"] / 100.0
            igtf_tasa = config_fin["igtf_pct"] / 100.0

            iva_usd = monto_base_usd * iva_tasa
            igtf_usd = (
                monto_base_usd * igtf_tasa
                if self._aplica_igtf(metodo_pago, config_fin)
                else 0.0
            )

            total_usd = monto_base_usd + iva_usd + igtf_usd
            total_bs = total_usd * tasa_bcv

            cursor_local.execute("""
                INSERT INTO solicitudes_servicio (
                    cliente_id, servicio_id, fecha_solicitud, monto,
                    iva_usd, igtf_usd, total_usd, tasa_bcv, total_bs,
                    metodo_pago, activo, estado_pago
                ) VALUES (?, ?, date('now'), ?, ?, ?, ?, ?, ?, ?, 'Activo', 'Pendiente')
            """, (
                cliente_id, servicio_id, monto_base_usd, iva_usd, igtf_usd,
                total_usd, tasa_bcv, total_bs, metodo_pago
            ))

            conn_local.commit()
            conn_local.close()

            # Marcar la solicitud web como aprobada en el servidor central.
            self._api_solicitudes(
                "PATCH",
                f"/api/solicitudes/{int(id_web_real)}",
                json={"estado": "Aprobado"}
            )

            usuario_actual = getattr(self, 'usuario_autenticado', 'Sistema')
            registrar_accion(usuario_actual, "APROBÓ SERVICIO WEB", f"Aprobó orden para {cliente_nombre}. Servicio: {servicio_interes}. Total: {total_bs:,.2f} Bs.")

            self._mensaje_corporativo(
                "Asesoría aprobada",
                f"{cliente_nombre} fue integrado correctamente como cliente.",
                tipo="exito",
                detalle=(
                    f"Servicio: {servicio_interes}\n"
                    f"Código cliente: {codigo_cliente}\n"
                    f"Tasa BCV: {tasa_bcv:.4f} Bs/USD\n"
                    f"Total: $ {total_usd:,.2f}  ·  Bs. {total_bs:,.2f}"
                )
            )
            self.mostrar_notificaciones()

        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo completar la operación",
                "Ocurrió un problema durante la aprobación de la asesoría.",
                tipo="error",
                detalle=str(e)
            )
            
    def mostrar_historial(self):

        if not self._requiere_administrador("consultar el historial completo del sistema"):
            return
        """Renderiza la pantalla de auditoría interna de la firma."""
        panel = self.crear_panel_principal("Historial de Actividades y Auditoría")
        
        tk.Label(panel, text="Registro cronológico inmutable de acciones del personal corporativo:", 
                bg=self.colores["fondo_main"], font=("Arial", 10, "italic")).pack(anchor="w", padx=25, pady=(0, 10))

        columnas_h = ("ID Registro", "Fecha y Hora", "Asesor / Usuario", "Operación Realizada", "Detalles del Movimiento")
        tabla_historial = ttk.Treeview(panel, columns=columnas_h, show="headings", height=18)
        
        for col in columnas_h:
            tabla_historial.heading(col, text=col)
            if col == "Detalles del Movimiento":
                tabla_historial.column(col, anchor="w", width=420)
            elif col == "Fecha y Hora":
                tabla_historial.column(col, anchor="center", width=160)
            else:
                tabla_historial.column(col, anchor="center", width=130)
                
        tabla_historial.pack(fill="both", expand=True, padx=25, pady=10)

        tk.Button(panel, text="🔄 Actualizar Bitácora", bg=self.colores["botones_menu"], fg=self.colores["oro"],
                font=("Arial", 10, "bold"), relief="flat", padx=10, pady=5,
                command=lambda: cargar_logs()).pack(anchor="w", padx=25, pady=5)

        def cargar_logs():
            for item in tabla_historial.get_children():
                tabla_historial.delete(item)
            try:
                conn = obtener_conexion()
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, datetime(fecha_hora, 'localtime'), usuario, accion, detalles 
                    FROM historial_actividades 
                    ORDER BY id DESC
                """)
                for fila in cursor.fetchall():
                    tabla_historial.insert("", "end", values=fila)
                conn.close()
            except Exception as e:
                print(f"Error cargando historial: {e}")

        cargar_logs()
        
        # ============================================================
    # ============================================================
    # MONITOR GLOBAL DE MENSAJES
    # ============================================================

    def iniciar_monitor_mensajes(self):
        """Inicia la vigilancia global de mensajes nuevos."""
        try:
            conn = obtener_conexion(); cursor = conn.cursor()
            cursor.execute("SELECT COALESCE(MAX(id), 0) FROM mensajes")
            resultado = cursor.fetchone()
            self.ultimo_mensaje_notificado = resultado[0] if resultado else 0
            conn.close()
        except Exception as e:
            print(f"Error iniciando monitor de mensajes: {e}")
            self.ultimo_mensaje_notificado = 0
        self.root.after(2000, self.monitor_mensajes)

    def monitor_mensajes(self):
        """Comprueba periódicamente si llegaron mensajes para el usuario actual."""
        try:
            usuario_actual = getattr(self, "usuario_autenticado", None)
            if not usuario_actual:
                self.root.after(2000, self.monitor_mensajes); return
            conn = obtener_conexion(); cursor = conn.cursor()
            cursor.execute("SELECT id FROM usuarios WHERE username = ?", (usuario_actual,))
            usuario = cursor.fetchone()
            if not usuario:
                conn.close(); self.root.after(2000, self.monitor_mensajes); return
            usuario_id = usuario[0]
            cursor.execute("""
                SELECT m.id, m.contenido, u.username, m.conversacion_id
                FROM mensajes m
                INNER JOIN usuarios u ON u.id = m.usuario_id
                INNER JOIN participantes_chat pc ON pc.conversacion_id = m.conversacion_id
                WHERE m.id > ? AND pc.usuario_id = ? AND m.usuario_id != ?
                  AND COALESCE(m.leido, 0) = 0
                ORDER BY m.id ASC
            """, (self.ultimo_mensaje_notificado, usuario_id, usuario_id))
            nuevos = cursor.fetchall(); conn.close()
            for mid, contenido, remitente, conv_id in nuevos:
                self.ultimo_mensaje_notificado = max(self.ultimo_mensaje_notificado, mid)
                self.mostrar_notificacion_chat(remitente, contenido or "", conv_id)
        except Exception as e:
            print(f"Error monitorizando mensajes: {e}")
        try:
            self.root.after(2000, self.monitor_mensajes)
        except Exception:
            pass

    def mostrar_notificacion_chat(self, remitente, contenido, conversacion_id):
        self.notificaciones_chat += 1
        n = tk.Toplevel(self.root); n.title("Nuevo mensaje"); n.resizable(False, False); n.configure(bg="white"); n.attributes("-topmost", True)
        ancho, alto = 350, 155; n.update_idletasks()
        sw, sh = n.winfo_screenwidth(), n.winfo_screenheight()
        desplazamiento = max(self.notificaciones_chat - 1, 0) * 12
        n.geometry(f"{ancho}x{alto}+{sw-ancho-25}+{max(20, sh-alto-70-desplazamiento)}")
        tk.Label(n, text="💬  Nuevo mensaje", bg="white", fg=self.colores["sidebar"], font=("Arial", 13, "bold")).pack(anchor="w", padx=15, pady=(12,3))
        tk.Label(n, text=remitente, bg="white", fg="#222222", font=("Arial", 10, "bold")).pack(anchor="w", padx=15)
        texto = contenido[:55] + "..." if len(contenido) > 55 else contenido
        tk.Label(n, text=texto, bg="white", fg="#666666", font=("Arial", 9), wraplength=315, justify="left").pack(anchor="w", padx=15, pady=(2,7))
        tk.Button(n, text="Abrir chat", bg=self.colores["sidebar"], fg="white", activebackground="#4f0e1c", activeforeground="white", relief="flat", cursor="hand2", command=lambda: self.abrir_chat_desde_notificacion(n, remitente, conversacion_id)).pack(padx=15, pady=(0,10), anchor="e")
        n.after(7000, lambda: self.cerrar_notificacion(n))

    def cerrar_notificacion(self, notificacion):
        try:
            if notificacion.winfo_exists(): return
            notificacion.destroy()
            if self.notificaciones_chat > 0: self.notificaciones_chat -= 1
        except Exception:
            pass

    def abrir_chat_desde_notificacion(self, notificacion, remitente, conversacion_id):
        self.cerrar_notificacion(notificacion)
        try:
            self.abrir_chat()
            self.root.after(300, lambda: self.seleccionar_usuario_chat(remitente))
        except Exception as e:
            self._mensaje_corporativo(
                "No se pudo abrir la conversación",
                "Ocurrió un problema al abrir el chat.",
                tipo="error",
                detalle=str(e)
            )

    def seleccionar_usuario_chat(self, username):
        try:
            ventana = getattr(self, "ventana_chat", None)
            if not ventana: return
            ventana.cargar_usuarios()
            for indice, datos in ventana.mapa_usuarios.items():
                usuario_id, nombre = datos
                if nombre == username:
                    ventana.lista_usuarios.selection_clear(0, tk.END)
                    ventana.lista_usuarios.selection_set(indice)
                    ventana.lista_usuarios.see(indice)
                    ventana.usuario_seleccionado()
                    return
        except Exception as e:
            print(f"Error seleccionando chat desde notificación: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = DashboardSGC(root)
    root.mainloop()